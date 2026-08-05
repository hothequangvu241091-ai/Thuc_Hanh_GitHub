# -*- coding: utf-8 -*-
"""V2: Nhập dữ liệu KE_HOACH, nhận diện trùng theo cùng file + Title [SEO] + H1."""

from __future__ import annotations

import os
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import win32com.client as win32
import win32api
import win32con
from openpyxl import load_workbook


SHEET_PLAN = "KE_HOACH"
KEY_HEADER = "Title [SEO]"
SOURCE_FILE_HEADER = "File nguồn"
SOURCE_LOCATION_HEADER = "Vị trí nguồn"
DOMAIN_HEADER = "Tên Miền"
URL_PAGE_HEADER = "URL Page"
DUPLICATE_MARKER = "Bài viết trùng"
NOTE_SAME_FILE_DIFFERENT_H1 = "trùng Title khác H1 cùng file"
NOTE_OTHER_FILE_SAME_H1 = "trùng Title và H1 file khác"
NOTE_OTHER_FILE_DIFFERENT_H1 = "trùng Title khác H1 file khác"
DOMAIN_SOURCE_ALIASES = ["Tên Miền", "Đợt viết"]
MAX_SOURCE_FILES = 50
MAX_SAFE_ROW = 5000
IMPORTED_ROW_HEIGHT = 15
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}

HEADER_GROUPS = [
    ([KEY_HEADER], KEY_HEADER),
    ([URL_PAGE_HEADER], URL_PAGE_HEADER),
    (DOMAIN_SOURCE_ALIASES, DOMAIN_HEADER),
    (["Main Keyword"], "Main Keyword"),
    (["Article Name [H1]"], "Article Name [H1]"),
    (["POST / UPDATE", "CATE [POST]"], "CATE [POST]"),
    (
        [
            """ĐẦU VÀO CỦA PROMPT
[Copy > Dán đúng loai Prompt]""",
            """ĐẦU VÀO CỦA PROMPT
[Copy > Dán đúng Prompt]""",
        ],
        """ĐẦU VÀO CỦA PROMPT
[Copy > Dán đúng Prompt]""",
    ),
    (
        [
            """CHÚ Ý ĐẶC BIỆT (!!!)
[PROMPT SỬ DỤNG]"""
        ],
        """CHÚ Ý ĐẶC BIỆT (!!!)
[PROMPT SỬ DỤNG]""",
    ),
    ([SOURCE_FILE_HEADER], SOURCE_FILE_HEADER),
    ([SOURCE_LOCATION_HEADER], SOURCE_LOCATION_HEADER),
]

INVALID_TEXT_VALUES = {
    "",
    "#",
    "-",
    "#.n/a",
    "#.na",
    "#n/a",
    "n/a",
    "na",
    "#value!",
    "#ref!",
    "#name?",
    "#div/0!",
    "#num!",
    "#null!",
}


def normalize(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    return text.casefold()


def domain_from_excel_filename(path: Path) -> str:
    """Lấy tên miền từ tên file Excel, ví dụ example.com.xlsx -> example.com."""
    name = path.stem.strip()
    if not re.fullmatch(
        r"(?:[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?\.)+[a-z]{2,63}",
        name,
        flags=re.IGNORECASE,
    ):
        return ""
    return name


def is_valid_value(value: Any) -> bool:
    if value is None:
        return False
    return normalize(value) not in INVALID_TEXT_VALUES


def is_valid_key(value: Any) -> bool:
    return is_valid_value(value)


def read_matrix(
    ws: Any,
    first_row: int,
    last_row: int,
    first_col: int,
    last_col: int,
) -> tuple:
    values = ws.Range(
        ws.Cells(first_row, first_col),
        ws.Cells(last_row, last_col),
    ).Value2
    if not isinstance(values, tuple):
        return ((values,),)
    if values and not isinstance(values[0], tuple):
        return (values,)
    return values


def headers(ws: Any) -> dict[str, tuple[int, str]]:
    last_col = int(ws.Cells(1, ws.Columns.Count).End(-4159).Column)
    result: dict[str, tuple[int, str]] = {}
    duplicates: list[str] = []
    for col in range(1, last_col + 1):
        raw = ws.Cells(1, col).Value2
        if raw is None or not str(raw).strip():
            continue
        key = normalize(raw)
        # Bỏ qua heading giữ chỗ hoặc lỗi như #, -, #N/A...
        if key in INVALID_TEXT_VALUES:
            continue
        if key in result:
            duplicates.append(str(raw).strip())
            continue
        result[key] = (col, str(raw).strip())
    if duplicates:
        raise RuntimeError(
            f"Sheet {ws.Name} có heading bị trùng: "
            + ", ".join(duplicates[:10])
        )
    return result


def last_value_row(ws: Any, col: int) -> int:
    found = ws.Columns(col).Find(
        What="*",
        After=ws.Cells(1, col),
        LookIn=-4163,       # xlValues
        LookAt=2,           # xlPart
        SearchOrder=1,      # xlByRows
        SearchDirection=2,  # xlPrevious
        MatchCase=False,
    )
    return 1 if found is None else int(found.Row)


def direct_headers(ws: Any) -> dict[str, tuple[int, str]]:
    result: dict[str, tuple[int, str]] = {}
    duplicates: list[str] = []
    required_source_keys = {
        normalize(name)
        for aliases, _canonical in HEADER_GROUPS
        for name in aliases
    }
    first_row = next(ws.iter_rows(min_row=1, max_row=1))
    for col, cell in enumerate(first_row, start=1):
        raw = cell.value
        if not is_valid_value(raw):
            continue
        raw_text = str(raw).strip()
        key = normalize(raw_text)
        # Chỉ quan tâm các heading thật sự cần nhập vào KE_HOACH.
        # Cột lạ hoặc dữ liệu thừa ở hàng 1 (ví dụ "272") bị bỏ qua.
        if key not in required_source_keys:
            continue
        if key in result:
            duplicates.append(raw_text)
            continue
        result[key] = (col, raw_text)
    if duplicates:
        raise RuntimeError(
            f"Sheet {ws.title} có heading bị trùng: "
            + ", ".join(duplicates[:10])
        )
    return result


def direct_last_value_row(ws: Any, col: int) -> int:
    # ReadOnlyWorksheet đọc tuần tự rất nhanh, nhưng gọi ws.cell() lặp ngược
    # có thể buộc openpyxl quét lại XML cho từng dòng. Vì vậy luôn quét cột
    # Title [SEO] đúng một lượt, đồng thời xử lý được cả max_row=None.
    last_row = 1
    for row_number, values in enumerate(
        ws.iter_rows(
            min_row=2,
            min_col=col,
            max_col=col,
            values_only=True,
        ),
        start=2,
    ):
        if is_valid_key(values[0]):
            last_row = row_number
    return last_row


def ask_source_mode(excel: Any) -> str | None:
    """Chọn nhập một file duy nhất hoặc quét toàn bộ thư mục."""
    result = win32api.MessageBox(
        int(excel.Hwnd),
        "Bạn muốn lấy dữ liệu theo cách nào?\n\n"
        "Yes: Chọn MỘT file Excel duy nhất\n"
        "No: Chọn THƯ MỤC chứa nhiều file Excel\n"
        "Cancel: Hủy chạy",
        "Chọn nguồn dữ liệu KE_HOACH",
        (
            win32con.MB_YESNOCANCEL
            | win32con.MB_ICONQUESTION
            | win32con.MB_DEFBUTTON1
        ),
    )
    if result == win32con.IDYES:
        return "file"
    if result == win32con.IDNO:
        return "folder"
    return None


def choose_file(excel: Any) -> str | None:
    """Chọn đúng một file Excel nguồn."""
    dialog = excel.FileDialog(3)  # msoFileDialogFilePicker
    dialog.Title = "Chọn một file Excel nguồn"
    dialog.AllowMultiSelect = False
    try:
        dialog.Filters.Clear()
        dialog.Filters.Add("File Excel", "*.xlsx;*.xlsm")
    except Exception:
        pass
    if dialog.Show() != -1:
        return None
    return str(dialog.SelectedItems.Item(1))


def choose_folder(excel: Any) -> str | None:
    dialog = excel.FileDialog(4)  # msoFileDialogFolderPicker
    dialog.Title = "Chọn thư mục chứa các file Excel nguồn"
    dialog.AllowMultiSelect = False
    if dialog.Show() != -1:
        return None
    return str(dialog.SelectedItems.Item(1))


def ask_include_subfolders(excel: Any) -> bool | None:
    result = win32api.MessageBox(
        int(excel.Hwnd),
        "Có lấy thêm các file Excel nằm trong thư mục con không?\n\n"
        "Yes: Có lấy thư mục con\n"
        "No: Chỉ lấy file nằm trực tiếp trong thư mục đã chọn\n"
        "Cancel: Hủy chạy",
        "Phạm vi lấy dữ liệu KE_HOACH",
        (
            win32con.MB_YESNOCANCEL
            | win32con.MB_ICONQUESTION
            | win32con.MB_DEFBUTTON2
        ),
    )
    if result == win32con.IDYES:
        return True
    if result == win32con.IDNO:
        return False
    return None


def excel_files(
    folder: str,
    target_full_name: str,
    include_subfolders: bool,
) -> list[Path]:
    target_path = os.path.normcase(os.path.abspath(target_full_name))
    root = Path(folder)
    result: list[Path] = []
    candidates = root.rglob("*") if include_subfolders else root.iterdir()
    for item in candidates:
        if not item.is_file():
            continue
        if item.name.startswith("~$"):
            continue
        if item.suffix.casefold() not in EXCEL_EXTENSIONS:
            continue
        if os.path.normcase(os.path.abspath(str(item))) == target_path:
            continue
        result.append(item)
    return sorted(
        result,
        key=lambda path: str(path.relative_to(root)).casefold(),
    )


def planned_target_headers(
    existing: dict[str, tuple[int, str]],
) -> tuple[dict[str, tuple[int, str]], list[tuple[int, str]]]:
    planned = dict(existing)
    additions: list[tuple[int, str]] = []
    next_col = max((item[0] for item in existing.values()), default=0) + 1

    for aliases, canonical in HEADER_GROUPS:
        canonical_key = normalize(canonical)
        if canonical_key in existing:
            selected = existing[canonical_key]
        else:
            existing_aliases = [
                existing[normalize(alias)]
                for alias in aliases
                if normalize(alias) in existing
            ]
            if existing_aliases:
                selected = existing_aliases[0]
            else:
                selected = (next_col, canonical)
                additions.append(selected)
                next_col += 1

        # Mọi tên đồng cấp ở nguồn đều trỏ về một cột đích duy nhất.
        planned[canonical_key] = selected
        for alias in aliases:
            planned[normalize(alias)] = selected
    return planned, additions


def build_existing_record_rows(
    ws: Any,
    target_headers: dict[str, tuple[int, str]],
) -> dict[tuple[str, str, str], int]:
    """Lập khóa ổn định: File nguồn + Title hiển thị + Article Name [H1].

    Nếu dữ liệu cũ đã có khóa trùng, giữ dòng xuất hiện đầu tiên và cảnh báo,
    không dừng toàn bộ quá trình nhập.
    """
    key_col = target_headers[normalize(KEY_HEADER)][0]
    h1_col = target_headers[normalize("Article Name [H1]")][0]
    source_file_col = target_headers[normalize(SOURCE_FILE_HEADER)][0]
    last_row = last_value_row(ws, key_col)
    if last_row < 2:
        return {}

    first_col = min(key_col, h1_col, source_file_col)
    last_col = max(key_col, h1_col, source_file_col)
    matrix = read_matrix(ws, 2, last_row, first_col, last_col)
    result: dict[tuple[str, str, str], int] = {}
    duplicates: list[str] = []
    for offset, row_values in enumerate(matrix, start=2):
        title = row_values[key_col - first_col]
        h1 = row_values[h1_col - first_col]
        source_file = row_values[source_file_col - first_col]
        if not (is_valid_key(title) and is_valid_value(h1) and is_valid_value(source_file)):
            continue
        record_key = (normalize(source_file), normalize(title), normalize(h1))
        if record_key in result:
            duplicates.append(
                f"dòng {offset}: {source_file} | {title} | {h1} "
                f"(đã có ở dòng {result[record_key]})"
            )
            continue
        result[record_key] = offset
    if duplicates:
        print(
            f"CẢNH BÁO: {SHEET_PLAN} đang có {len(duplicates)} khóa "
            "File nguồn + Title + H1 bị trùng; giữ dòng đầu tiên."
        )
        for item in duplicates[:20]:
            print("- " + item)
    return result


def _append_duplicate_note(title: str, note: str) -> str:
    """Thêm hậu tố phân loại, tránh nhân hậu tố khi chạy lại dữ liệu đã ghi chú."""
    clean = str(title or "").strip()
    suffix = f" - {note}"
    if normalize(clean).endswith(normalize(suffix)):
        return clean
    return clean + suffix


def collect_source_records(
    files: list[Path],
    root_folder: str,
    target_headers: dict[str, tuple[int, str]],
) -> tuple[
    dict[tuple[str, str, str], tuple[dict[int, Any], str, str]],
    list[str],
    set[str],
    int,
    list[str],
    set[tuple[str, str, str]],
    list[tuple[str, str, int, int]],
]:
    records: dict[tuple[str, str, str], tuple[dict[int, Any], str, str]] = {}
    skipped_workbooks: list[str] = []
    ignored_headers: set[str] = set()
    skipped_invalid_rows = 0
    duplicate_notes: list[str] = []
    duplicate_highlight_keys: set[tuple[str, str, str]] = set()
    duplicate_source_marks: list[tuple[str, str, int, int]] = []
    root = Path(root_folder)

    # Theo dõi Title gốc trên toàn bộ đợt nhập để nhận biết trùng giữa các file.
    # Mỗi phần tử: (file_key, h1_key, source_ref).
    seen_title_global: dict[str, list[tuple[str, str, str]]] = defaultdict(list)

    for path in files:
        source_file_label = str(path.relative_to(root))
        source_file_key = normalize(source_file_label)
        source_filename_domain = domain_from_excel_filename(path)
        source_book = None
        found_source_sheet = False

        # Chỉ dùng để loại dòng trùng hoàn toàn trong cùng một file.
        seen_exact_in_file: dict[tuple[str, str], str] = {}
        try:
            source_book = load_workbook(filename=path, read_only=True, data_only=True)
            for ws in source_book.worksheets:
                source_header_map = direct_headers(ws)
                key_info = source_header_map.get(normalize(KEY_HEADER))
                if key_info is None:
                    continue
                found_source_sheet = True
                key_col = key_info[0]
                keyword_info = source_header_map.get(normalize("Main Keyword"))
                h1_info = source_header_map.get(normalize("Article Name [H1]"))
                if keyword_info is None or h1_info is None:
                    missing_combo_headers = []
                    if keyword_info is None:
                        missing_combo_headers.append("Main Keyword")
                    if h1_info is None:
                        missing_combo_headers.append("Article Name [H1]")
                    print(
                        f"Bỏ qua sheet: {source_file_label} / {ws.title} vì thiếu cột bắt buộc: "
                        + ", ".join(missing_combo_headers)
                    )
                    continue
                keyword_col = keyword_info[0]
                h1_col = h1_info[0]
                url_info = source_header_map.get(normalize(URL_PAGE_HEADER))
                url_col = url_info[0] if url_info is not None else 0
                print(f"Đang đọc nguồn: {source_file_label} / {ws.title}")
                last_row = direct_last_value_row(ws, key_col)
                if last_row < 2:
                    continue

                last_col = max(item[0] for item in source_header_map.values())
                matrix = tuple(tuple(row) for row in ws.iter_rows(
                    min_row=2, max_row=last_row, min_col=1, max_col=last_col, values_only=True
                ))
                source_to_target: dict[int, int] = {}
                for source_name, (source_col, raw_name) in source_header_map.items():
                    target_info = target_headers.get(source_name)
                    if target_info is None:
                        ignored_headers.add(raw_name)
                        continue
                    source_to_target[source_col] = target_info[0]

                source_ref_prefix = f"{source_file_label} / {ws.title}"
                has_domain_heading = any(normalize(name) in source_header_map for name in DOMAIN_SOURCE_ALIASES)
                if not has_domain_heading:
                    if key_col == 1:
                        raise RuntimeError(
                            f"{source_ref_prefix}: thiếu cả '{DOMAIN_HEADER}' và 'Đợt viết', "
                            f"nhưng cột A đang là '{KEY_HEADER}', nên không thể dùng cột A làm Tên Miền an toàn."
                        )
                    source_to_target[1] = target_headers[normalize(DOMAIN_HEADER)][0]

                for offset, row_values in enumerate(matrix):
                    source_row = offset + 2
                    raw_title = row_values[key_col - 1]
                    raw_keyword = row_values[keyword_col - 1]
                    raw_h1 = row_values[h1_col - 1]
                    if not (is_valid_key(raw_title) and is_valid_value(raw_keyword) and is_valid_value(raw_h1)):
                        skipped_invalid_rows += 1
                        continue

                    original_title = str(raw_title).strip()
                    h1_text = str(raw_h1).strip()
                    title_key = normalize(original_title)
                    h1_key = normalize(h1_text)
                    exact_key = (title_key, h1_key)
                    source_location = f"{ws.title} - dòng {source_row}"
                    source_ref = f"{source_file_label} / {source_location}"

                    # 1) Cùng file + cùng Title + cùng H1: không nhập dòng thứ hai.
                    # Chỉ ghi dấu vào URL Page của dòng nguồn nếu ô còn trống.
                    if exact_key in seen_exact_in_file:
                        previous_ref = seen_exact_in_file[exact_key]
                        if url_col:
                            duplicate_source_marks.append(
                                (str(path), ws.title, source_row, url_col)
                            )
                        else:
                            duplicate_notes.append(
                                f"'{original_title}' + H1 '{h1_text}' tại {source_ref} "
                                f"trùng hoàn toàn với {previous_ref}; không thể ghi "
                                f"'{DUPLICATE_MARKER}' vì thiếu cột {URL_PAGE_HEADER}."
                            )
                        duplicate_notes.append(
                            f"BỎ QUA: '{original_title}' + H1 '{h1_text}' tại "
                            f"{source_ref} trùng hoàn toàn trong cùng file với {previous_ref}."
                        )
                        continue

                    adjusted_title = original_title
                    note = ""
                    prior_entries = seen_title_global.get(title_key, [])

                    # Phân loại dòng vẫn được nhập. Ưu tiên trùng chính xác ở file khác,
                    # tiếp theo là cùng file nhưng H1 khác, cuối cùng là file khác H1 khác.
                    other_file_same_h1 = next(
                        (entry for entry in prior_entries
                         if entry[0] != source_file_key and entry[1] == h1_key),
                        None,
                    )
                    same_file_different_h1 = next(
                        (entry for entry in prior_entries
                         if entry[0] == source_file_key and entry[1] != h1_key),
                        None,
                    )
                    other_file_different_h1 = next(
                        (entry for entry in prior_entries
                         if entry[0] != source_file_key and entry[1] != h1_key),
                        None,
                    )

                    matched_ref = ""
                    if other_file_same_h1 is not None:
                        note = NOTE_OTHER_FILE_SAME_H1
                        matched_ref = other_file_same_h1[2]
                    elif same_file_different_h1 is not None:
                        note = NOTE_SAME_FILE_DIFFERENT_H1
                        matched_ref = same_file_different_h1[2]
                    elif other_file_different_h1 is not None:
                        note = NOTE_OTHER_FILE_DIFFERENT_H1
                        matched_ref = other_file_different_h1[2]

                    if note:
                        adjusted_title = _append_duplicate_note(original_title, note)
                        duplicate_notes.append(
                            f"GHI CHÚ: '{original_title}' tại {source_ref} thành "
                            f"'{adjusted_title}' vì đối chiếu với {matched_ref}."
                        )

                    # Ghi nhận bản đầu tiên hợp lệ của cặp Title + H1 trong file.
                    seen_exact_in_file[exact_key] = source_ref
                    seen_title_global[title_key].append(
                        (source_file_key, h1_key, source_ref)
                    )

                    record_key = (
                        source_file_key,
                        normalize(adjusted_title),
                        h1_key,
                    )
                    if note:
                        duplicate_highlight_keys.add(record_key)

                    values_by_target_col: dict[int, Any] = {}
                    for source_col, target_col in source_to_target.items():
                        value = row_values[source_col - 1]
                        if not is_valid_value(value):
                            continue
                        previous_value = values_by_target_col.get(target_col)
                        if previous_value is not None and normalize(previous_value) != normalize(value):
                            raise RuntimeError(
                                f"{source_ref} có hai heading đồng cấp nhưng dữ liệu khác nhau: "
                                f"'{str(previous_value).strip()}' và '{str(value).strip()}'."
                            )
                        values_by_target_col[target_col] = value
                    values_by_target_col[target_headers[normalize(KEY_HEADER)][0]] = adjusted_title
                    if source_filename_domain:
                        values_by_target_col[target_headers[normalize(DOMAIN_HEADER)][0]] = source_filename_domain
                    records[record_key] = (
                        values_by_target_col,
                        source_file_label,
                        source_location,
                    )
        finally:
            if source_book is not None:
                source_book.close()

        if not found_source_sheet:
            skipped_workbooks.append(source_file_label)

    return (
        records,
        skipped_workbooks,
        ignored_headers,
        skipped_invalid_rows,
        duplicate_notes,
        duplicate_highlight_keys,
        duplicate_source_marks,
    )

def mark_duplicates_in_source_files(
    excel: Any,
    marks: list[tuple[str, str, int, int]],
) -> tuple[int, int, list[str]]:
    """Ghi 'Bài viết trùng' vào URL Page của dòng trùng; không ghi đè ô đã có dữ liệu."""
    written = 0
    preserved = 0
    errors: list[str] = []
    grouped: dict[str, list[tuple[str, int, int]]] = defaultdict(list)
    for file_path, sheet_name, row, col in marks:
        grouped[file_path].append((sheet_name, row, col))

    for file_path, file_marks in grouped.items():
        workbook = None
        opened_here = False
        try:
            normalized = os.path.normcase(os.path.abspath(file_path))
            for wb in excel.Workbooks:
                try:
                    if os.path.normcase(os.path.abspath(str(wb.FullName))) == normalized:
                        workbook = wb
                        break
                except Exception:
                    continue
            if workbook is None:
                workbook = excel.Workbooks.Open(file_path, UpdateLinks=0, ReadOnly=False)
                opened_here = True

            changed = False
            for sheet_name, row, col in file_marks:
                ws = workbook.Worksheets(sheet_name)
                cell = ws.Cells(row, col)
                current = cell.Value2
                if not is_valid_value(current):
                    cell.Value2 = DUPLICATE_MARKER
                    written += 1
                    changed = True
                else:
                    preserved += 1
            if changed:
                workbook.Save()
        except Exception as exc:
            errors.append(f"{file_path}: {type(exc).__name__}: {exc}")
        finally:
            if opened_here and workbook is not None:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    pass
    return written, preserved, errors


def contiguous_runs(row_values: dict[int, Any]) -> list[list[tuple[int, Any]]]:
    items = sorted(row_values.items())
    if not items:
        return []
    runs: list[list[tuple[int, Any]]] = [[items[0]]]
    for item in items[1:]:
        if item[0] == runs[-1][-1][0] + 1:
            runs[-1].append(item)
        else:
            runs.append([item])
    return runs


def main() -> None:
    print("=" * 72)
    print("NHẬP DỮ LIỆU KE_HOACH V2.1 - PHÂN LOẠI BÀI TRÙNG")
    print("=" * 72)

    excel = win32.GetActiveObject("Excel.Application")
    target_book = None
    ws_plan = None
    for workbook in excel.Workbooks:
        try:
            candidate = workbook.Worksheets(SHEET_PLAN)
            target_book = workbook
            ws_plan = candidate
            break
        except Exception:
            continue
    if target_book is None or ws_plan is None:
        raise RuntimeError(
            f"Không tìm thấy workbook Excel đang mở nào có sheet {SHEET_PLAN}."
        )
    print(f"Workbook đích: {target_book.Name}")

    source_mode = ask_source_mode(excel)
    if source_mode is None:
        print("Đã hủy chọn nguồn. Chưa có dữ liệu nào được thay đổi.")
        return

    if source_mode == "file":
        selected_file = choose_file(excel)
        if selected_file is None:
            print("Đã hủy chọn file. Chưa có dữ liệu nào được thay đổi.")
            return

        source_path = Path(selected_file).resolve()
        target_path = Path(str(target_book.FullName)).resolve()
        if source_path.suffix.casefold() not in EXCEL_EXTENSIONS:
            raise RuntimeError("File đã chọn không phải định dạng .xlsx hoặc .xlsm.")
        if os.path.normcase(str(source_path)) == os.path.normcase(str(target_path)):
            raise RuntimeError("Không được chọn chính workbook đích làm file nguồn.")

        folder = str(source_path.parent)
        files = [source_path]
        print(f"Phạm vi: chỉ một file nguồn: {source_path.name}")
    else:
        folder = choose_folder(excel)
        if folder is None:
            print("Đã hủy chọn thư mục. Chưa có dữ liệu nào được thay đổi.")
            return

        include_subfolders = ask_include_subfolders(excel)
        if include_subfolders is None:
            print("Đã hủy chọn phạm vi. Chưa có dữ liệu nào được thay đổi.")
            return
        print(
            "Phạm vi: "
            + (
                "thư mục đã chọn và toàn bộ thư mục con."
                if include_subfolders
                else "chỉ file nằm trực tiếp trong thư mục đã chọn."
            )
        )

        files = excel_files(
            folder,
            target_book.FullName,
            include_subfolders,
        )
        if not files:
            raise RuntimeError("Thư mục đã chọn không có file Excel nguồn phù hợp.")
    if len(files) > MAX_SOURCE_FILES:
        preview = ", ".join(
            str(path.relative_to(Path(folder)))
            for path in files[:10]
        )
        raise RuntimeError(
            f"Thư mục đã chọn có {len(files)} file Excel, vượt giới hạn "
            f"an toàn {MAX_SOURCE_FILES} file. Có thể bạn đã chọn nhầm "
            f"thư mục. Chương trình chưa mở hay ghi dữ liệu. "
            f"Ví dụ file tìm thấy: {preview}"
        )

    existing_headers = headers(ws_plan)
    target_headers, heading_additions = planned_target_headers(existing_headers)
    key_col = target_headers[normalize(KEY_HEADER)][0]
    existing_record_rows = build_existing_record_rows(ws_plan, target_headers)

    (
        source_records,
        skipped_workbooks,
        ignored_headers,
        skipped_invalid_rows,
        duplicate_notes,
        duplicate_highlight_keys,
        duplicate_source_marks,
    ) = collect_source_records(files, folder, target_headers)

    if not source_records:
        raise RuntimeError(
            "Không tìm thấy dòng nào có Title [SEO] hợp lệ. "
            "Chương trình chưa thay đổi KE_HOACH."
        )

    next_row = max(
        last_value_row(ws_plan, key_col) + 1,
        2,
    )
    updates_by_col: dict[int, dict[int, Any]] = defaultdict(dict)
    added_rows = 0
    existing_rows_seen = 0
    existing_urls_added = 0
    changed_cells = 0
    duplicate_title_rows: list[int] = []

    # Đọc giá trị hiện có một lần cho từng cột để chỉ ghi những ô thật sự đổi.
    last_existing_row = max(next_row - 1, 1)
    current_values_by_col: dict[int, dict[int, Any]] = {}
    used_target_cols = sorted(
        {
            col
            for values, _source_file, _source_location in source_records.values()
            for col in values
        }
        | {
            target_headers[normalize(SOURCE_FILE_HEADER)][0],
            target_headers[normalize(SOURCE_LOCATION_HEADER)][0],
        }
    )
    for col in used_target_cols:
        current: dict[int, Any] = {}
        if last_existing_row >= 2:
            matrix = read_matrix(ws_plan, 2, last_existing_row, col, col)
            current = {
                offset + 2: item[0]
                for offset, item in enumerate(matrix)
            }
        current_values_by_col[col] = current

    source_file_col = target_headers[normalize(SOURCE_FILE_HEADER)][0]
    source_location_col = target_headers[normalize(SOURCE_LOCATION_HEADER)][0]
    url_page_col = target_headers[normalize(URL_PAGE_HEADER)][0]

    for record_key, (
        source_values,
        source_file,
        source_location,
    ) in source_records.items():
        target_row = existing_record_rows.get(record_key)
        if target_row is not None:
            existing_rows_seen += 1
            if record_key in duplicate_highlight_keys:
                duplicate_title_rows.append(target_row)
            # Dòng cũ chỉ được bổ sung URL Page còn trống; không cập nhật
            # hoặc ghi đè bất kỳ dữ liệu nào khác.
            source_url = source_values.get(url_page_col)
            current_url = current_values_by_col.get(
                url_page_col, {}
            ).get(target_row)
            if (
                is_valid_value(source_url)
                and not is_valid_value(current_url)
            ):
                updates_by_col[url_page_col][target_row] = source_url
                changed_cells += 1
                existing_urls_added += 1
            continue

        target_row = next_row
        next_row += 1
        added_rows += 1
        row_changed = True
        if record_key in duplicate_highlight_keys:
            duplicate_title_rows.append(target_row)

        for col, value in source_values.items():
            old_value = current_values_by_col.get(col, {}).get(target_row)
            if normalize(old_value) == normalize(value):
                continue
            updates_by_col[col][target_row] = value
            changed_cells += 1
            row_changed = True

        # Chỉ đổi thông tin truy vết khi dòng mới hoặc dữ liệu thật sự đổi.
        if row_changed:
            for col, value in (
                (source_file_col, source_file),
                (source_location_col, source_location),
            ):
                old_value = current_values_by_col.get(col, {}).get(target_row)
                if normalize(old_value) == normalize(value):
                    continue
                updates_by_col[col][target_row] = value
                changed_cells += 1

    final_last_row = next_row - 1
    if final_last_row > MAX_SAFE_ROW:
        raise RuntimeError(
            f"Sau khi nhập, {SHEET_PLAN} sẽ tới dòng {final_last_row:,}, "
            f"vượt giới hạn an toàn {MAX_SAFE_ROW:,}. Chưa ghi dữ liệu."
        )

    old_screen_updating = excel.ScreenUpdating
    old_enable_events = excel.EnableEvents
    old_calculation = excel.Calculation
    try:
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        excel.Calculation = -4135  # xlCalculationManual

        for col, heading in heading_additions:
            ws_plan.Cells(1, col).Value2 = heading

        affected_rows = {
            row
            for row_values in updates_by_col.values()
            for row in row_values
        }
        if affected_rows:
            first_write_row = min(affected_rows)
            last_write_row = max(affected_rows)
            first_write_col = min(updates_by_col)
            last_write_col = max(updates_by_col)
            write_matrix = [
                list(row)
                for row in read_matrix(
                    ws_plan,
                    first_write_row,
                    last_write_row,
                    first_write_col,
                    last_write_col,
                )
            ]
            for col, row_values in updates_by_col.items():
                for row, value in row_values.items():
                    write_matrix[row - first_write_row][
                        col - first_write_col
                    ] = value
            ws_plan.Range(
                ws_plan.Cells(first_write_row, first_write_col),
                ws_plan.Cells(last_write_row, last_write_col),
            ).Value2 = tuple(tuple(row) for row in write_matrix)

        # Thu gọn cả dữ liệu cũ: các Title đã có sẽ bị bỏ qua ở lần chạy sau,
        # nhưng chiều cao hàng của chúng vẫn phải được sửa.
        if final_last_row >= 2:
            last_used_col = max(item[0] for item in target_headers.values())
            compact_range = ws_plan.Range(
                ws_plan.Cells(2, 1),
                ws_plan.Cells(final_last_row, last_used_col),
            )
            compact_range.WrapText = False
            compact_range.Rows.RowHeight = IMPORTED_ROW_HEIGHT

        for row in sorted(set(duplicate_title_rows)):
            ws_plan.Cells(row, key_col).Interior.Color = 255

        target_book.Save()
    finally:
        excel.Calculation = old_calculation
        excel.EnableEvents = old_enable_events
        excel.ScreenUpdating = old_screen_updating

    duplicate_marks_written, duplicate_marks_preserved, duplicate_mark_errors = (
        mark_duplicates_in_source_files(excel, duplicate_source_marks)
    )

    print(f"Đã quét: {len(files)} file Excel.")
    print(f"Đã thêm heading còn thiếu: {len(heading_additions)}.")
    print(f"Đã thêm dòng mới: {added_rows}.")
    print(
        f"Title [SEO] đã có trong KE_HOACH: "
        f"{existing_rows_seen}."
    )
    print(
        "Đã bổ sung URL Page vào dòng cũ còn trống: "
        f"{existing_urls_added}."
    )
    print(
        f"Đã xử lý các trường hợp trùng/nghi trùng: "
        f"{len(duplicate_notes)}."
    )
    print(f"Đã ghi '{DUPLICATE_MARKER}' vào URL nguồn: {duplicate_marks_written}.")
    print(f"Ô URL nguồn đã có dữ liệu nên giữ nguyên: {duplicate_marks_preserved}.")
    if duplicate_mark_errors:
        print("Lỗi khi ghi dấu bài trùng về nguồn:")
        for item in duplicate_mark_errors[:20]:
            print("- " + item)
    print(f"Tổng số ô được cập nhật: {changed_cells}.")
    print(
        "Dòng nguồn bỏ qua vì thiếu bộ ba "
        "Title [SEO] + Article Name [H1] + Main Keyword: "
        f"{skipped_invalid_rows}."
    )
    if skipped_workbooks:
        print(
            "File không có sheet chứa Title [SEO], đã bỏ qua: "
            + ", ".join(skipped_workbooks)
        )
    if ignored_headers:
        print(
            "Heading nguồn chưa có trong KE_HOACH, đã bỏ qua: "
            + " | ".join(sorted(ignored_headers, key=str.casefold))
        )
    if duplicate_notes:
        print("Chi tiết xử lý bài trùng và nghi trùng:")
        for item in duplicate_notes[:30]:
            print("- " + item)
    print("Hoàn tất. Dữ liệu và các dòng cũ không có trong nguồn được giữ nguyên.")


if __name__ == "__main__":
    main()

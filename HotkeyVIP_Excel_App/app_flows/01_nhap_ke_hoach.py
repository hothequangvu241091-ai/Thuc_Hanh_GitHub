# -*- coding: utf-8 -*-
"""V5.5: Chỉ đọc sheet Article; chống trùng và đồng bộ lại dòng đã có.

Khi một bản ghi đã tồn tại trong KE_HOACH theo bộ Title + H1 + Main Keyword
+ File nguồn, toàn bộ cột dữ liệu có trong sheet Article được đồng bộ vào dòng
hiện có, kể cả giá trị trống. Giữ nguyên File nguồn và Trạng thái nguồn; cập nhật
Vị trí nguồn theo số dòng hiện tại trong sheet Article.
"""

from __future__ import annotations

import os
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import win32com.client as win32
import win32api
import win32con
from openpyxl import load_workbook


APP_WORKBOOK: Any = None
SHEET_PLAN = "KE_HOACH"
KEY_HEADER = "Title [SEO]"
SOURCE_FILE_HEADER = "File nguồn"
SOURCE_LOCATION_HEADER = "Vị trí nguồn"
SOURCE_STATUS_HEADER = "Trạng thái nguồn"
RecordIdentity = tuple[str, str, int]
DOMAIN_HEADER = "Tên Miền"
URL_PAGE_HEADER = "URL Page"
DOMAIN_SOURCE_ALIASES = ["Tên Miền", "Đợt viết"]
MAX_SOURCE_FILES = 50
MAX_SAFE_ROW = 20000
IMPORTED_ROW_HEIGHT = 15
DUPLICATE_FILL_COLOR = win32api.RGB(255, 199, 206)
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
CONTROL_WORKBOOK_NAME = "1.QL_New-Content_Kênh SEO.xlsx"
CONTROL_CHANNEL_HEADER = "NHÓM KÊNH"
CONTROL_ANALYSIS_HEADER = "PHÂN TÍCH"
CONTROL_ANALYSIS_COMPLETE = "Phân tích xong"
DEFAULT_SOURCE_FOLDER = (
    r"G:\.shortcut-targets-by-id\1Emi7P7uNkYpOjfn6wOnl9VeQbJesvcwP"
    r"\1.New-Content_Kênh SEO"
)

HEADER_GROUPS = [
    ([KEY_HEADER], KEY_HEADER),
    (["Search Question"], "Search Question"),
    ([URL_PAGE_HEADER], URL_PAGE_HEADER),
    (DOMAIN_SOURCE_ALIASES, DOMAIN_HEADER),
    (["Main Keyword"], "Main Keyword"),
    (["Article Name [H1]"], "Article Name [H1]"),
    (["POST / UPDATE", "CATE [POST]"], "CATE [POST]"),
    (
        [
            """ĐẦU VÀO CỦA PROMPT
[Copy > Dán đúng loai Prompt]""",
            """ĐẦU VÀO CỦA PROMPT
[Copy > Dán đúng Prompt]""",
        ],
        """ĐẦU VÀO CỦA PROMPT
[Copy > Dán đúng Prompt]""",
    ),
    (
        [
            """CHÚ Ý ĐẶC BIỆT (!!!)
[PROMPT SỬ DỤNG]"""
        ],
        """CHÚ Ý ĐẶC BIỆT (!!!)
[PROMPT SỬ DỤNG]""",
    ),
    ([SOURCE_FILE_HEADER], SOURCE_FILE_HEADER),
    ([SOURCE_LOCATION_HEADER], SOURCE_LOCATION_HEADER),
    ([SOURCE_STATUS_HEADER], SOURCE_STATUS_HEADER),
]

INVALID_TEXT_VALUES = {
    "",
    "#",
    "-",
    "#.n/a",
    "#.na",
    "#n/a",
    "n/a",
    "na",
    "#value!",
    "#ref!",
    "#name?",
    "#div/0!",
    "#num!",
    "#null!",
}


def normalize(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    return text.casefold()


def domain_from_excel_filename(path: Path) -> str:
    """Lấy tên miền từ tên file Excel, ví dụ example.com.xlsx -> example.com."""
    name = path.stem.strip()
    if not re.fullmatch(
        r"(?:[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?\.)+[a-z]{2,63}",
        name,
        flags=re.IGNORECASE,
    ):
        return ""
    return name


def is_valid_value(value: Any) -> bool:
    if value is None:
        return False
    return normalize(value) not in INVALID_TEXT_VALUES


def is_valid_key(value: Any) -> bool:
    return is_valid_value(value)


def read_matrix(
    ws: Any,
    first_row: int,
    last_row: int,
    first_col: int,
    last_col: int,
) -> tuple:
    values = ws.Range(
        ws.Cells(first_row, first_col),
        ws.Cells(last_row, last_col),
    ).Value2
    if not isinstance(values, tuple):
        return ((values,),)
    if values and not isinstance(values[0], tuple):
        return (values,)
    return values


def headers(ws: Any) -> dict[str, tuple[int, str]]:
    last_col = int(ws.Cells(1, ws.Columns.Count).End(-4159).Column)
    result: dict[str, tuple[int, str]] = {}
    duplicates: list[str] = []
    for col in range(1, last_col + 1):
        raw = ws.Cells(1, col).Value2
        if raw is None or not str(raw).strip():
            continue
        key = normalize(raw)
        # Bỏ qua heading giữ chỗ hoặc lỗi như #, -, #N/A...
        if key in INVALID_TEXT_VALUES:
            continue
        if key in result:
            duplicates.append(str(raw).strip())
            continue
        result[key] = (col, str(raw).strip())
    if duplicates:
        raise RuntimeError(
            f"Sheet {ws.Name} có heading bị trùng: "
            + ", ".join(duplicates[:10])
        )
    return result


def direct_headers(ws: Any) -> dict[str, tuple[int, str]]:
    result: dict[str, tuple[int, str]] = {}
    duplicates: list[str] = []
    required_source_keys = {
        normalize(name)
        for aliases, _canonical in HEADER_GROUPS
        for name in aliases
    }
    first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if first_row is None:
        print(f"Bỏ qua sheet trống: {ws.title}")
        return result
    for col, cell in enumerate(first_row, start=1):
        raw = cell.value
        if not is_valid_value(raw):
            continue
        raw_text = str(raw).strip()
        key = normalize(raw_text)
        # Chỉ quan tâm các heading thật sự cần nhập vào KE_HOACH.
        # Cột lạ hoặc dữ liệu thừa ở hàng 1 (ví dụ "272") bị bỏ qua.
        if key not in required_source_keys:
            continue
        if key in result:
            duplicates.append(raw_text)
            continue
        result[key] = (col, raw_text)
    if duplicates:
        raise RuntimeError(
            f"Sheet {ws.title} có heading bị trùng: "
            + ", ".join(duplicates[:10])
        )
    return result


def direct_last_value_row(ws: Any, col: int) -> int:
    # ReadOnlyWorksheet đọc tuần tự rất nhanh, nhưng gọi ws.cell() lặp ngược
    # có thể buộc openpyxl quét lại XML cho từng dòng. Vì vậy luôn quét cột
    # Title [SEO] đúng một lượt, đồng thời xử lý được cả max_row=None.
    last_row = 1
    for row_number, values in enumerate(
        ws.iter_rows(
            min_row=2,
            min_col=col,
            max_col=col,
            values_only=True,
        ),
        start=2,
    ):
        if is_valid_key(values[0]):
            last_row = row_number
    return last_row


def last_data_row(ws: Any, col: int) -> int:
    """Tìm dòng cuối có giá trị thật trong một cột Excel COM."""
    found = ws.Columns(col).Find(
        What="*", After=ws.Cells(1, col), LookIn=-4163, LookAt=2,
        SearchOrder=1, SearchDirection=2, MatchCase=False,
    )
    return 1 if found is None else int(found.Row)


def delete_rows_in_batches(ws: Any, rows: list[int]) -> int:
    """Xóa đúng từng dòng từ dưới lên để không bị lệch số dòng."""
    if not rows:
        return 0

    sorted_rows = sorted(set(rows))
    for row in reversed(sorted_rows):
        ws.Rows(row).Delete()
    return len(sorted_rows)




def ask_run_mode(excel: Any) -> str | None:
    """Chọn chế độ mặc định tự động hoặc cách chọn nguồn thủ công cũ."""
    result = win32api.MessageBox(
        int(excel.Hwnd),
        "Bạn muốn chạy theo chế độ nào?\n\n"
        "Yes: MẶC ĐỊNH - tự đọc file tổng và chỉ lấy kênh đã phân tích xong\n"
        "No: CHỌN TAY - chọn một file hoặc thư mục như cách cũ\n"
        "Cancel: Hủy chạy",
        "Chọn chế độ nhập KE_HOACH",
        (
            win32con.MB_YESNOCANCEL
            | win32con.MB_ICONQUESTION
            | win32con.MB_DEFBUTTON1
        ),
    )
    if result == win32con.IDYES:
        return "default"
    if result == win32con.IDNO:
        return "manual"
    return None


def ask_source_mode(excel: Any) -> str | None:
    """Chọn nhập một file duy nhất hoặc quét toàn bộ thư mục."""
    result = win32api.MessageBox(
        int(excel.Hwnd),
        "Bạn muốn lấy dữ liệu theo cách nào?\n\n"
        "Yes: Chọn MỘT file Excel duy nhất\n"
        "No: Chọn THƯ MỤC chứa nhiều file Excel\n"
        "Cancel: Hủy chạy",
        "Chọn nguồn dữ liệu KE_HOACH",
        (
            win32con.MB_YESNOCANCEL
            | win32con.MB_ICONQUESTION
            | win32con.MB_DEFBUTTON1
        ),
    )
    if result == win32con.IDYES:
        return "file"
    if result == win32con.IDNO:
        return "folder"
    return None


def choose_file(excel: Any) -> str | None:
    """Chọn đúng một file Excel nguồn."""
    dialog = excel.FileDialog(3)  # msoFileDialogFilePicker
    dialog.Title = "Chọn một file Excel nguồn"
    dialog.AllowMultiSelect = False
    try:
        dialog.Filters.Clear()
        dialog.Filters.Add("File Excel", "*.xlsx;*.xlsm")
    except Exception:
        pass
    if dialog.Show() != -1:
        return None
    return str(dialog.SelectedItems.Item(1))


def choose_folder(excel: Any) -> str | None:
    dialog = excel.FileDialog(4)  # msoFileDialogFolderPicker
    dialog.Title = "Chọn thư mục chứa các file Excel nguồn"
    dialog.AllowMultiSelect = False
    if dialog.Show() != -1:
        return None
    return str(dialog.SelectedItems.Item(1))


def ask_include_subfolders(excel: Any) -> bool | None:
    result = win32api.MessageBox(
        int(excel.Hwnd),
        "Có lấy thêm các file Excel nằm trong thư mục con không?\n\n"
        "Yes: Có lấy thư mục con\n"
        "No: Chỉ lấy file nằm trực tiếp trong thư mục đã chọn\n"
        "Cancel: Hủy chạy",
        "Phạm vi lấy dữ liệu KE_HOACH",
        (
            win32con.MB_YESNOCANCEL
            | win32con.MB_ICONQUESTION
            | win32con.MB_DEFBUTTON2
        ),
    )
    if result == win32con.IDYES:
        return True
    if result == win32con.IDNO:
        return False
    return None


def excel_files(
    folder: str,
    target_full_name: str,
    include_subfolders: bool,
) -> list[Path]:
    target_path = os.path.normcase(os.path.abspath(target_full_name))
    root = Path(folder)
    result: list[Path] = []
    candidates = root.rglob("*") if include_subfolders else root.iterdir()
    for item in candidates:
        if not item.is_file():
            continue
        if item.name.startswith("~$"):
            continue
        if item.suffix.casefold() not in EXCEL_EXTENSIONS:
            continue
        if os.path.normcase(os.path.abspath(str(item))) == target_path:
            continue
        result.append(item)
    return sorted(
        result,
        key=lambda path: str(path.relative_to(root)).casefold(),
    )


def planned_target_headers(
    existing: dict[str, tuple[int, str]],
) -> tuple[dict[str, tuple[int, str]], list[tuple[int, str]]]:
    planned = dict(existing)
    additions: list[tuple[int, str]] = []
    next_col = max((item[0] for item in existing.values()), default=0) + 1

    for aliases, canonical in HEADER_GROUPS:
        canonical_key = normalize(canonical)
        if canonical_key in existing:
            selected = existing[canonical_key]
        else:
            existing_aliases = [
                existing[normalize(alias)]
                for alias in aliases
                if normalize(alias) in existing
            ]
            if existing_aliases:
                selected = existing_aliases[0]
            else:
                selected = (next_col, canonical)
                additions.append(selected)
                next_col += 1

        # Mọi tên đồng cấp ở nguồn đều trỏ về một cột đích duy nhất.
        planned[canonical_key] = selected
        for alias in aliases:
            planned[normalize(alias)] = selected
    return planned, additions


def collect_source_records(
    files: list[Path],
    root_folder: str,
    target_headers: dict[str, tuple[int, str]],
) -> tuple[
    list[tuple[dict[int, Any], str, str, RecordIdentity]],
    list[str],
    set[str],
    int,
]:
    """Thu thập mọi dòng có đủ Title, Main Keyword và H1."""
    records: list[tuple[dict[int, Any], str, str, RecordIdentity]] = []
    skipped_workbooks: list[str] = []
    ignored_headers: set[str] = set()
    skipped_invalid_rows = 0
    root = Path(root_folder)

    for path in files:
        source_file_label = str(path.relative_to(root))
        source_filename_domain = domain_from_excel_filename(path)
        source_book = None
        found_source_sheet = False
        try:
            source_book = load_workbook(filename=path, read_only=True, data_only=True)

            # V5.3: chỉ đọc đúng sheet "Article", bỏ qua toàn bộ sheet khác
            # như "Lọc", "Sheet1", v.v.
            article_ws = None
            for candidate_ws in source_book.worksheets:
                if normalize(candidate_ws.title) == normalize("Article"):
                    article_ws = candidate_ws
                    break

            if article_ws is None:
                print(f"Bỏ qua file không có sheet Article: {source_file_label}")
            else:
                ws = article_ws
                source_header_map = direct_headers(ws)
                key_info = source_header_map.get(normalize(KEY_HEADER))
                if key_info is None:
                    print(
                        f"Bỏ qua sheet Article: {source_file_label} vì thiếu cột '{KEY_HEADER}'."
                    )
                else:
                    found_source_sheet = True
                    key_col = key_info[0]
                    keyword_info = source_header_map.get(normalize("Main Keyword"))
                    h1_info = source_header_map.get(normalize("Article Name [H1]"))
                    if keyword_info is None or h1_info is None:
                        missing_combo_headers = []
                        if keyword_info is None:
                            missing_combo_headers.append("Main Keyword")
                        if h1_info is None:
                            missing_combo_headers.append("Article Name [H1]")
                        print(
                            f"Bỏ qua sheet: {source_file_label} / {ws.title} vì thiếu cột bắt buộc: "
                            + ", ".join(missing_combo_headers)
                        )
                        continue
                    keyword_col = keyword_info[0]
                    h1_col = h1_info[0]
                    print(f"Đang đọc nguồn: {source_file_label} / {ws.title}")
                    last_row = direct_last_value_row(ws, key_col)
                    if last_row < 2:
                        continue

                    last_col = max(item[0] for item in source_header_map.values())
                    matrix = tuple(tuple(row) for row in ws.iter_rows(
                        min_row=2, max_row=last_row, min_col=1, max_col=last_col, values_only=True
                    ))
                    source_to_target: dict[int, int] = {}
                    for source_name, (source_col, raw_name) in source_header_map.items():
                        # Cột này do chương trình quản lý trong KE_HOACH, không
                        # nhận giá trị từ file nguồn.
                        if source_name == normalize(SOURCE_STATUS_HEADER):
                            continue
                        target_info = target_headers.get(source_name)
                        if target_info is None:
                            ignored_headers.add(raw_name)
                            continue
                        source_to_target[source_col] = target_info[0]

                    source_ref_prefix = f"{source_file_label} / {ws.title}"
                    has_domain_heading = any(normalize(name) in source_header_map for name in DOMAIN_SOURCE_ALIASES)
                    if not has_domain_heading:
                        if key_col == 1:
                            raise RuntimeError(
                                f"{source_ref_prefix}: thiếu cả '{DOMAIN_HEADER}' và 'Đợt viết', "
                                f"nhưng cột A đang là '{KEY_HEADER}', nên không thể dùng cột A làm Tên Miền an toàn."
                            )
                        source_to_target[1] = target_headers[normalize(DOMAIN_HEADER)][0]

                    for offset, row_values in enumerate(matrix):
                        source_row = offset + 2
                        raw_title = row_values[key_col - 1]
                        raw_keyword = row_values[keyword_col - 1]
                        raw_h1 = row_values[h1_col - 1]
                        if not (is_valid_key(raw_title) and is_valid_value(raw_keyword) and is_valid_value(raw_h1)):
                            skipped_invalid_rows += 1
                            continue

                        original_title = str(raw_title).strip()
                        h1_text = str(raw_h1).strip()
                        source_location = f"{ws.title} - dòng {source_row}"
                        record_key: RecordIdentity = (
                            source_file_label,
                            ws.title,
                            source_row,
                        )

                        values_by_target_col: dict[int, Any] = {}
                        for source_col, target_col in source_to_target.items():
                            value = row_values[source_col - 1]
                            previous_value = values_by_target_col.get(target_col)
                            if (
                                target_col in values_by_target_col
                                and is_valid_value(previous_value)
                                and is_valid_value(value)
                                and normalize(previous_value) != normalize(value)
                            ):
                                raise RuntimeError(
                                    f"{source_ref_prefix} - dòng {source_row} có hai heading "
                                    "đồng cấp nhưng dữ liệu khác nhau: "
                                    f"'{str(previous_value).strip()}' và '{str(value).strip()}'."
                                )
                            # Nếu hai heading đồng cấp cùng trỏ về một cột, ưu tiên
                            # giá trị thật; nếu cả hai trống thì vẫn giữ ô trống để
                            # đồng bộ ghi đè được dữ liệu cũ trong KE_HOACH.
                            if (
                                target_col not in values_by_target_col
                                or not is_valid_value(previous_value)
                                or is_valid_value(value)
                            ):
                                values_by_target_col[target_col] = value
                        values_by_target_col[target_headers[normalize(KEY_HEADER)][0]] = original_title
                        if source_filename_domain:
                            values_by_target_col[target_headers[normalize(DOMAIN_HEADER)][0]] = source_filename_domain
                        records.append((
                            values_by_target_col,
                            source_file_label,
                            source_location,
                            record_key,
                        ))
        finally:
            if source_book is not None:
                source_book.close()

        if not found_source_sheet:
            skipped_workbooks.append(source_file_label)

    return (
        records,
        skipped_workbooks,
        ignored_headers,
        skipped_invalid_rows,
    )


def clear_active_filters(ws: Any) -> int:
    """Bỏ mọi bộ lọc đang ẩn dòng trong KE_HOACH trước khi làm mới dữ liệu."""
    cleared = 0
    try:
        if bool(ws.FilterMode):
            ws.ShowAllData()
            cleared += 1
    except Exception:
        pass

    try:
        for index in range(1, int(ws.ListObjects.Count) + 1):
            table_filter = ws.ListObjects.Item(index).AutoFilter
            if bool(table_filter.FilterMode):
                table_filter.ShowAllData()
                cleared += 1
    except Exception:
        pass
    return cleared


def clear_all_plan_data(ws: Any) -> int:
    """Xóa toàn bộ dữ liệu KE_HOACH từ dòng 2, giữ lại header và định dạng."""
    last_by_row = ws.Cells.Find(
        What="*",
        After=ws.Cells(1, 1),
        LookIn=-4123,       # xlFormulas
        LookAt=2,           # xlPart
        SearchOrder=1,      # xlByRows
        SearchDirection=2,  # xlPrevious
        MatchCase=False,
    )
    last_by_col = ws.Cells.Find(
        What="*",
        After=ws.Cells(1, 1),
        LookIn=-4123,
        LookAt=2,
        SearchOrder=2,      # xlByColumns
        SearchDirection=2,
        MatchCase=False,
    )
    if last_by_row is None or last_by_col is None or int(last_by_row.Row) < 2:
        return 0

    last_row = int(last_by_row.Row)
    last_col = int(last_by_col.Column)
    ws.Range(ws.Cells(2, 1), ws.Cells(last_row, last_col)).ClearContents()
    return last_row - 1


def reset_plan_before_rebuild(excel: Any, ws: Any) -> tuple[int, int]:
    """Ẩn màn hình, bỏ filter và clear KE_HOACH trước khi ghi dữ liệu mới."""
    old_screen_updating = excel.ScreenUpdating
    old_enable_events = excel.EnableEvents
    try:
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        filters_cleared = clear_active_filters(ws)
        removed_rows = clear_all_plan_data(ws)
    finally:
        excel.EnableEvents = old_enable_events
        excel.ScreenUpdating = old_screen_updating
    return removed_rows, filters_cleared


def find_control_workbook(folder: str) -> Path | None:
    """Tìm file tổng ở ngay folder đã chọn, không quét vào thư mục con."""
    wanted_name = CONTROL_WORKBOOK_NAME.casefold()
    for item in Path(folder).iterdir():
        if item.is_file() and item.name.casefold() == wanted_name:
            return item
    return None


def control_headers(ws: Any) -> dict[str, int]:
    result: dict[str, int] = {}
    for col, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1):
        if not is_valid_value(cell.value):
            continue
        header_key = normalize(cell.value)
        if header_key in result:
            raise RuntimeError(f"Sheet {ws.title} có heading bị trùng: {cell.value}")
        result[header_key] = col
    return result


def ready_domains_from_control_file(path: Path) -> tuple[set[str], list[str]]:
    """Đọc tên miền tại các dòng có PHÂN TÍCH là 'Phân tích xong'."""
    domains: set[str] = set()
    invalid_domains: list[str] = []
    found_sheet = False
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        for ws in workbook.worksheets:
            header_map = control_headers(ws)
            channel_col = header_map.get(normalize(CONTROL_CHANNEL_HEADER))
            analysis_col = header_map.get(normalize(CONTROL_ANALYSIS_HEADER))
            if channel_col is None or analysis_col is None:
                continue
            found_sheet = True
            last_col = max(channel_col, analysis_col)
            for row_values in ws.iter_rows(
                min_row=2,
                min_col=1,
                max_col=last_col,
                values_only=True,
            ):
                if normalize(row_values[analysis_col - 1]) != normalize(CONTROL_ANALYSIS_COMPLETE):
                    continue
                raw_domain = row_values[channel_col - 1]
                if not is_valid_value(raw_domain):
                    continue
                domain = str(raw_domain).strip()
                if not domain_from_excel_filename(Path(domain + ".xlsx")):
                    invalid_domains.append(domain)
                    continue
                domains.add(normalize(domain))
    finally:
        workbook.close()

    if not found_sheet:
        raise RuntimeError(
            f"{path.name} không có sheet chứa đủ heading "
            f"'{CONTROL_CHANNEL_HEADER}' và '{CONTROL_ANALYSIS_HEADER}'."
        )
    return domains, invalid_domains


def select_ready_domain_files(
    folder: str,
    target_full_name: str,
    control_path: Path,
    include_subfolders: bool,
    ready_domains: set[str],
) -> tuple[list[Path], list[str], list[str]]:
    """Chọn đúng file Excel có tên bằng tên miền đã phân tích xong."""
    root = Path(folder)
    target_path = os.path.normcase(os.path.abspath(target_full_name))
    control_full_path = os.path.normcase(os.path.abspath(str(control_path)))
    candidates = root.rglob("*") if include_subfolders else root.iterdir()
    matches_by_domain: dict[str, list[Path]] = defaultdict(list)
    for item in candidates:
        if not item.is_file() or item.name.startswith("~$"):
            continue
        if item.suffix.casefold() not in EXCEL_EXTENSIONS:
            continue
        item_path = os.path.normcase(os.path.abspath(str(item)))
        if item_path in {target_path, control_full_path}:
            continue
        domain_key = normalize(item.stem)
        if domain_key in ready_domains:
            matches_by_domain[domain_key].append(item)

    selected: list[Path] = []
    missing: list[str] = []
    ambiguous: list[str] = []
    for domain in sorted(ready_domains):
        matches = matches_by_domain.get(domain, [])
        if len(matches) == 1:
            selected.append(matches[0])
        elif not matches:
            missing.append(domain)
        else:
            ambiguous.append(domain)
    return (
        sorted(selected, key=lambda item: str(item.relative_to(root)).casefold()),
        missing,
        ambiguous,
    )

def main() -> None:
    print("=" * 72)
    print("NHẬP KE_HOACH V5.3 - CHỈ ĐỌC SHEET ARTICLE + CHỐNG TRÙNG COMBO 3 + FILE NGUỒN")
    print("=" * 72)

    if APP_WORKBOOK is not None:
        target_book = APP_WORKBOOK
        excel = target_book.Application
        ws_plan = target_book.Worksheets(SHEET_PLAN)
    else:
        excel = win32.GetActiveObject("Excel.Application")
        target_book = None
        ws_plan = None
        for workbook in excel.Workbooks:
            try:
                candidate = workbook.Worksheets(SHEET_PLAN)
                target_book = workbook
                ws_plan = candidate
                break
            except Exception:
                continue
    if target_book is None or ws_plan is None:
        raise RuntimeError(
            f"Không tìm thấy workbook Excel đang mở nào có sheet {SHEET_PLAN}."
        )
    print(f"Workbook đích: {target_book.Name}")

    # Kiểm tra quyền ghi trước khi quét nguồn hay thay đổi bất kỳ dữ liệu nào.
    # Thuộc tính Locked của ô chỉ được Excel thi hành khi sheet đang Protect.
    if bool(target_book.ReadOnly):
        raise RuntimeError(
            f"Workbook đích '{target_book.Name}' đang mở chỉ-đọc. "
            "Không thể xóa hoặc ghi dữ liệu vào KE_HOACH."
        )
    if bool(ws_plan.ProtectContents):
        raise RuntimeError(
            f"Sheet {SHEET_PLAN} đang được Protect. Hãy bỏ bảo vệ sheet "
            "trước khi chạy; chương trình chưa thay đổi dữ liệu."
        )

    run_mode = ask_run_mode(excel)
    if run_mode is None:
        print("Đã hủy chạy. Chưa có dữ liệu nào được thay đổi.")
        return

    auto_control_mode = run_mode == "default"
    ready_domain_names: set[str] | None = None
    control_missing_files: list[str] = []
    control_ambiguous_files: list[str] = []
    control_invalid_domains: list[str] = []

    if auto_control_mode:
        source_mode = "folder"
        folder = DEFAULT_SOURCE_FOLDER
        include_subfolders = False
        root_folder = Path(folder)
        if not root_folder.is_dir():
            raise RuntimeError(
                "Không tìm thấy folder mặc định:\n" + DEFAULT_SOURCE_FOLDER
            )
        control_path = find_control_workbook(folder)
        if control_path is None:
            raise RuntimeError(
                "Không tìm thấy file tổng '" + CONTROL_WORKBOOK_NAME
                + "' trong folder mặc định:\n" + DEFAULT_SOURCE_FOLDER
            )
        ready_domain_names, control_invalid_domains = ready_domains_from_control_file(
            control_path
        )
        if not ready_domain_names:
            print(
                "File tổng không có kênh nào có PHÂN TÍCH = "
                f"'{CONTROL_ANALYSIS_COMPLETE}'. Chưa có dữ liệu nào được thay đổi."
            )
            return
        files, control_missing_files, control_ambiguous_files = select_ready_domain_files(
            folder,
            target_book.FullName,
            control_path,
            include_subfolders,
            ready_domain_names,
        )
        print(f"Chế độ mặc định: {control_path.name}")
        print(f"Kênh đã phân tích xong: {len(ready_domain_names)}.")
        print(f"Tìm thấy file nguồn để xử lý: {len(files)}.")
    else:
        source_mode = ask_source_mode(excel)
        if source_mode is None:
            print("Đã hủy chọn nguồn. Chưa có dữ liệu nào được thay đổi.")
            return

        if source_mode == "file":
            selected_file = choose_file(excel)
            if selected_file is None:
                print("Đã hủy chọn file. Chưa có dữ liệu nào được thay đổi.")
                return

            source_path = Path(selected_file).resolve()
            target_path = Path(str(target_book.FullName)).resolve()
            if source_path.suffix.casefold() not in EXCEL_EXTENSIONS:
                raise RuntimeError("File đã chọn không phải định dạng .xlsx hoặc .xlsm.")
            if os.path.normcase(str(source_path)) == os.path.normcase(str(target_path)):
                raise RuntimeError("Không được chọn chính workbook đích làm file nguồn.")

            folder = str(source_path.parent)
            files = [source_path]
            include_subfolders = False
            print(f"Phạm vi: chỉ một file nguồn: {source_path.name}")
        else:
            folder = choose_folder(excel)
            if folder is None:
                print("Đã hủy chọn thư mục. Chưa có dữ liệu nào được thay đổi.")
                return

            include_subfolders = ask_include_subfolders(excel)
            if include_subfolders is None:
                print("Đã hủy chọn phạm vi. Chưa có dữ liệu nào được thay đổi.")
                return
            print(
                "Phạm vi: "
                + (
                    "thư mục đã chọn và toàn bộ thư mục con."
                    if include_subfolders
                    else "chỉ file nằm trực tiếp trong thư mục đã chọn."
                )
            )

            files = excel_files(
                folder,
                target_book.FullName,
                include_subfolders,
            )
            if not files:
                raise RuntimeError("Thư mục đã chọn không có file Excel nguồn phù hợp.")
    if len(files) > MAX_SOURCE_FILES:
        preview = ", ".join(
            str(path.relative_to(Path(folder)))
            for path in files[:10]
        )
        raise RuntimeError(
            f"Thư mục đã chọn có {len(files)} file Excel, vượt giới hạn "
            f"an toàn {MAX_SOURCE_FILES} file. Có thể bạn đã chọn nhầm "
            f"thư mục. Chương trình chưa mở hay ghi dữ liệu. "
            f"Ví dụ file tìm thấy: {preview}"
        )

    existing_headers = headers(ws_plan)
    target_headers, heading_additions = planned_target_headers(existing_headers)
    key_col = target_headers[normalize(KEY_HEADER)][0]
    h1_col = target_headers[normalize("Article Name [H1]")][0]
    keyword_col = target_headers[normalize("Main Keyword")][0]
    source_file_col = target_headers[normalize(SOURCE_FILE_HEADER)][0]
    source_location_col = target_headers[normalize(SOURCE_LOCATION_HEADER)][0]
    (
        source_records,
        skipped_workbooks,
        ignored_headers,
        skipped_invalid_rows,
    ) = collect_source_records(files, folder, target_headers)

    existing_last_row = last_data_row(ws_plan, key_col)
    first_data_col = min(item[0] for item in target_headers.values())
    last_data_col = max(item[0] for item in target_headers.values())
    existing_values_by_row: dict[int, dict[int, Any]] = {}
    existing_target_combos: set[tuple[str, str, str, str]] = set()
    existing_rows_by_combo: dict[tuple[str, str, str, str], list[int]] = defaultdict(list)
    if existing_last_row >= 2:
        existing_matrix = read_matrix(
            ws_plan, 2, existing_last_row, first_data_col, last_data_col
        )
        for target_row, values in enumerate(existing_matrix, start=2):
            row_values = {
                col: values[col - first_data_col]
                for col in range(first_data_col, last_data_col + 1)
            }
            existing_values_by_row[target_row] = row_values
            combo = (
                normalize(row_values.get(key_col)),
                normalize(row_values.get(h1_col)),
                normalize(row_values.get(keyword_col)),
                normalize(row_values.get(source_file_col)),
            )
            # Chỉ dùng làm khóa chống trùng khi xác định được cả 3 nội dung
            # và File nguồn. Khác File nguồn => được xem là bản ghi khác.
            if all(combo):
                existing_target_combos.add(combo)
                existing_rows_by_combo[combo].append(target_row)

    # V5.2: chống trùng theo COMBO 4:
    # Title [SEO] + Article Name [H1] + Main Keyword + File nguồn.
    # Cùng 3 nội dung nhưng KHÁC File nguồn => vẫn giữ cả hai.
    # Chỉ khi cùng 3 nội dung VÀ cùng File nguồn => giữ lần đầu, bỏ lần sau.
    seen_combos = set(existing_target_combos)
    filtered_source_records = []
    skipped_duplicate_source_combos = 0
    existing_row_updates: dict[int, dict[int, Any]] = {}
    refreshed_existing_combos: set[tuple[str, str, str, str]] = set()
    for record in source_records:
        source_values, source_file, source_location, _record_key = record
        combo = (
            normalize(source_values.get(key_col)),
            normalize(source_values.get(h1_col)),
            normalize(source_values.get(keyword_col)),
            normalize(source_file),
        )
        if combo in seen_combos:
            # V5.5: chỉ lấy bản ghi nguồn đầu tiên nếu nguồn tự có combo trùng.
            # Với combo đã có trong KE_HOACH, đồng bộ mọi cột đọc được từ Article,
            # kể cả ô trống. File nguồn là một phần của khóa nên luôn giữ nguyên;
            # Trạng thái nguồn đã bị loại ngay lúc đọc; Vị trí nguồn luôn làm mới.
            if (
                combo in existing_target_combos
                and combo not in refreshed_existing_combos
            ):
                incoming_values = dict(source_values)
                incoming_values.pop(source_file_col, None)
                incoming_values[source_location_col] = source_location
                normalized_updates = {
                    col: "" if value is None else value
                    for col, value in incoming_values.items()
                }
                for target_row in existing_rows_by_combo.get(combo, []):
                    existing_row_updates[target_row] = dict(normalized_updates)
                refreshed_existing_combos.add(combo)
            skipped_duplicate_source_combos += 1
            continue
        seen_combos.add(combo)
        filtered_source_records.append(record)

    source_records = filtered_source_records

    if not source_records and not existing_row_updates:
        print("KE_HOACH đã có đủ combo 3 + File nguồn; không cần thêm dòng mới.")
        return

    predicted_last_row = existing_last_row + len(source_records)
    if predicted_last_row > MAX_SAFE_ROW:
        raise RuntimeError(
            f"Sau khi nhập, {SHEET_PLAN} sẽ tới dòng {predicted_last_row:,}, "
            f"vượt giới hạn an toàn {MAX_SAFE_ROW:,}. KE_HOACH chưa thay đổi."
        )

    # Giữ nguyên KE_HOACH hiện có; chỉ thêm combo 3 + File nguồn chưa tồn tại. File nguồn chỉ được đọc.
    cleared_filters = clear_active_filters(ws_plan)

    next_row = max(existing_last_row + 1, 2)
    updates_by_col: dict[int, dict[int, Any]] = defaultdict(dict)
    added_rows = 0
    synchronized_cells = 0
    updated_existing_rows = len(existing_row_updates)
    updated_existing_cells = sum(len(values) for values in existing_row_updates.values())
    changed_cells = updated_existing_cells

    for target_row, incoming_values in existing_row_updates.items():
        for col, value in incoming_values.items():
            updates_by_col[col][target_row] = value
            existing_values_by_row[target_row][col] = value

    for source_values, source_file, source_location, _record_key in source_records:
        incoming_values = dict(source_values)
        incoming_values[source_file_col] = source_file
        incoming_values[source_location_col] = source_location

        target_row = next_row
        next_row += 1
        added_rows += 1

        current_values = {
            col: "" for col in range(first_data_col, last_data_col + 1)
        }
        for col, value in incoming_values.items():
            excel_value = "" if value is None else value
            updates_by_col[col][target_row] = excel_value
            current_values[col] = excel_value
            changed_cells += 1
            synchronized_cells += 1
        existing_values_by_row[target_row] = current_values

    final_last_row = max(existing_last_row, next_row - 1)
    old_screen_updating = excel.ScreenUpdating
    old_enable_events = excel.EnableEvents
    old_calculation = excel.Calculation
    try:
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        excel.Calculation = -4135  # xlCalculationManual

        for col, heading in heading_additions:
            ws_plan.Cells(1, col).Value2 = heading

        affected_rows = {
            row
            for row_values in updates_by_col.values()
            for row in row_values
        }
        if affected_rows:
            first_write_row = min(affected_rows)
            last_write_row = max(affected_rows)
            first_write_col = min(updates_by_col)
            last_write_col = max(updates_by_col)
            write_matrix = [
                list(row)
                for row in read_matrix(
                    ws_plan,
                    first_write_row,
                    last_write_row,
                    first_write_col,
                    last_write_col,
                )
            ]
            for col, row_values in updates_by_col.items():
                for row, value in row_values.items():
                    write_matrix[row - first_write_row][
                        col - first_write_col
                    ] = value
            ws_plan.Range(
                ws_plan.Cells(first_write_row, first_write_col),
                ws_plan.Cells(last_write_row, last_write_col),
            ).Value2 = tuple(tuple(row) for row in write_matrix)

        # Xóa màu cảnh báo cũ rồi thu gọn toàn bộ vùng dữ liệu sau khi nhập.
        last_used_col = max(item[0] for item in target_headers.values())
        last_styled_row = final_last_row
        if last_styled_row >= 2:
            ws_plan.Range(
                ws_plan.Cells(2, 1),
                ws_plan.Cells(last_styled_row, last_used_col),
            ).Interior.ColorIndex = -4142  # xlColorIndexNone
        if final_last_row >= 2:
            compact_range = ws_plan.Range(
                ws_plan.Cells(2, 1),
                ws_plan.Cells(final_last_row, last_used_col),
            )
            compact_range.WrapText = False
            compact_range.Rows.RowHeight = IMPORTED_ROW_HEIGHT

        target_book.Save()
    finally:
        excel.Calculation = old_calculation
        excel.EnableEvents = old_enable_events
        excel.ScreenUpdating = old_screen_updating

    print(f"Đã quét: {len(files)} file Excel.")
    print(f"Đã bỏ filter đang ẩn dòng: {cleared_filters}.")
    print("KE_HOACH được giữ nguyên; không xóa dữ liệu cũ.")
    print(f"Đã thêm heading còn thiếu: {len(heading_additions)}.")
    print(f"Đã thêm dòng mới: {added_rows}.")
    print(f"Đã bỏ bản ghi trùng đủ Title + H1 + Main Keyword + File nguồn: {skipped_duplicate_source_combos}.")
    print(f"Đã đồng bộ lại dòng KE_HOACH trùng combo 4: {updated_existing_rows}.")
    print(f"Số ô đã đồng bộ trên dòng hiện có: {updated_existing_cells}.")
    print(f"Số ô đã nhập mới từ nguồn: {synchronized_cells}.")
    print(f"Tổng số ô được cập nhật: {changed_cells}.")
    print(
        "Dòng nguồn bỏ qua vì thiếu bộ ba "
        "Title [SEO] + Article Name [H1] + Main Keyword: "
        f"{skipped_invalid_rows}."
    )
    if skipped_workbooks:
        print(
            "File không có sheet chứa Title [SEO], đã bỏ qua: "
            + ", ".join(skipped_workbooks)
        )
    if control_missing_files:
        print(
            "Kênh đã phân tích xong nhưng chưa có file Excel cùng tên: "
            + " | ".join(control_missing_files)
        )
    if control_ambiguous_files:
        print(
            "Kênh có nhiều file Excel trùng tên, đã bỏ qua để tránh chọn nhầm: "
            + " | ".join(control_ambiguous_files)
        )
    if control_invalid_domains:
        print(
            "NHÓM KÊNH không phải tên miền hợp lệ, đã bỏ qua: "
            + " | ".join(control_invalid_domains)
        )
    if ignored_headers:
        print(
            "Heading nguồn chưa có trong KE_HOACH, đã bỏ qua: "
            + " | ".join(sorted(ignored_headers, key=str.casefold))
        )
    print(
        "Hoàn tất. Chỉ đọc sheet Article. Khác File nguồn vẫn được giữ; "
        "bản ghi trùng combo 4 không thêm lại mà được đồng bộ từ nguồn."
    )


if __name__ == "__main__":
    main()

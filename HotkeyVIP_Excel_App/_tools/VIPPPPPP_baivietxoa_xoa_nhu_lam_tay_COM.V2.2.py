# -*- coding: utf-8 -*-
"""
V2.2 - XÓA "BÀI ĐÃ XÓA" NHƯ THAO TÁC TAY TRONG EXCEL

Cách chọn file:
- Đọc file tổng 1.QL_New-Content_Kênh SEO.xlsx.
- Chỉ lấy NHÓM KÊNH có PHÂN TÍCH = "Phân tích xong".
- Chỉ tìm file Excel cùng tên domain ngay trong thư mục gốc.
- KHÔNG quét thư mục con.
- Không mở 47 file bừa như V2.1.

Cách xóa:
- Dùng Microsoft Excel thật qua COM.
- Bỏ toàn bộ Filter trên sheet Article trước khi xóa.
- Tìm URL Page = "Bài đã xóa" hoặc "Bài đã xóa (đợt 2)".
- Xóa nguyên hàng từ dưới lên bằng EntireRow.Delete().
- Excel tự cập nhật công thức như thao tác tay.
- CalculateFullRebuild() rồi Save.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import os
import re
from typing import Any

import win32com.client as win32
from openpyxl import load_workbook


# ============================================================
# CẤU HÌNH
# ============================================================

ROOT_FOLDER = Path(
    r"G:\.shortcut-targets-by-id\1Emi7P7uNkYpOjfn6wOnl9VeQbJesvcwP\1.New-Content_Kênh SEO"
)

CONTROL_WORKBOOK_NAME = "1.QL_New-Content_Kênh SEO.xlsx"
CONTROL_CHANNEL_HEADER = "NHÓM KÊNH"
CONTROL_ANALYSIS_HEADER = "PHÂN TÍCH"
CONTROL_ANALYSIS_COMPLETE = "Phân tích xong"

TARGET_SHEET = "Article"
TARGET_COLUMN = "URL Page"

DELETE_VALUES = {
    "bài đã xóa",
    "bài đã xóa (đợt 2)",
}

EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
SHEET_PASSWORD = "11"


# ============================================================
# CHUẨN HÓA
# ============================================================

def normalize(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    return text.casefold()


def is_valid_value(value: Any) -> bool:
    return bool(normalize(value))


def domain_from_excel_filename(path: Path) -> str:
    """Kiểm tra tên có phải dạng domain hợp lệ không."""
    name = path.stem.strip()
    if not re.fullmatch(
        r"(?:[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?\.)+[a-z]{2,63}",
        name,
        flags=re.IGNORECASE,
    ):
        return ""
    return name


# ============================================================
# ĐỌC FILE TỔNG - GIỐNG V5.2
# ============================================================

def control_headers(ws: Any) -> dict[str, int]:
    result: dict[str, int] = {}
    first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if first_row is None:
        return result

    for col, cell in enumerate(first_row, start=1):
        if not is_valid_value(cell.value):
            continue

        key = normalize(cell.value)
        if key in result:
            raise RuntimeError(
                f"Sheet {ws.title} có heading bị trùng: {cell.value}"
            )

        result[key] = col

    return result


def ready_domains_from_control_file(path: Path) -> tuple[set[str], list[str]]:
    """
    Lấy NHÓM KÊNH ở những dòng có:
    PHÂN TÍCH = "Phân tích xong"
    """
    domains: set[str] = set()
    invalid_domains: list[str] = []
    found_sheet = False

    wb = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        for ws in wb.worksheets:
            header_map = control_headers(ws)

            channel_col = header_map.get(normalize(CONTROL_CHANNEL_HEADER))
            analysis_col = header_map.get(normalize(CONTROL_ANALYSIS_HEADER))

            if channel_col is None or analysis_col is None:
                continue

            found_sheet = True
            last_col = max(channel_col, analysis_col)

            for row_values in ws.iter_rows(
                min_row=2,
                min_col=1,
                max_col=last_col,
                values_only=True,
            ):
                if normalize(row_values[analysis_col - 1]) != normalize(
                    CONTROL_ANALYSIS_COMPLETE
                ):
                    continue

                raw_domain = row_values[channel_col - 1]

                if not is_valid_value(raw_domain):
                    continue

                domain = str(raw_domain).strip()

                if not domain_from_excel_filename(Path(domain + ".xlsx")):
                    invalid_domains.append(domain)
                    continue

                domains.add(normalize(domain))

    finally:
        wb.close()

    if not found_sheet:
        raise RuntimeError(
            f"{path.name} không có sheet chứa đủ heading "
            f"'{CONTROL_CHANNEL_HEADER}' và '{CONTROL_ANALYSIS_HEADER}'."
        )

    return domains, invalid_domains


def select_ready_domain_files(
    folder: Path,
    control_path: Path,
    ready_domains: set[str],
) -> tuple[list[Path], list[str], list[str]]:
    """
    Chỉ tìm file ngay trong ROOT_FOLDER.
    KHÔNG quét thư mục con.
    File được chọn khi tên file = domain trong file tổng.
    """
    control_full_path = os.path.normcase(
        os.path.abspath(str(control_path))
    )

    matches_by_domain: dict[str, list[Path]] = defaultdict(list)

    # CHỈ THƯ MỤC GỐC - KHÔNG RGlob.
    for item in folder.iterdir():
        if not item.is_file():
            continue

        if item.name.startswith("~$"):
            continue

        if item.suffix.casefold() not in EXCEL_EXTENSIONS:
            continue

        item_path = os.path.normcase(
            os.path.abspath(str(item))
        )

        if item_path == control_full_path:
            continue

        domain_key = normalize(item.stem)

        if domain_key in ready_domains:
            matches_by_domain[domain_key].append(item)

    selected: list[Path] = []
    missing: list[str] = []
    ambiguous: list[str] = []

    for domain in sorted(ready_domains):
        matches = matches_by_domain.get(domain, [])

        if len(matches) == 1:
            selected.append(matches[0])
        elif not matches:
            missing.append(domain)
        else:
            ambiguous.append(domain)

    selected.sort(key=lambda p: p.name.casefold())

    return selected, missing, ambiguous


# ============================================================
# FILTER
# ============================================================

def clear_all_filters(ws: Any) -> int:
    """Bỏ filter thường và filter trong Excel Table."""
    cleared = 0

    try:
        if bool(ws.FilterMode):
            ws.ShowAllData()
            cleared += 1
    except Exception:
        pass

    try:
        table_count = int(ws.ListObjects.Count)
    except Exception:
        table_count = 0

    for i in range(1, table_count + 1):
        try:
            af = ws.ListObjects.Item(i).AutoFilter
            if bool(af.FilterMode):
                af.ShowAllData()
                cleared += 1
        except Exception:
            pass

    return cleared


# ============================================================
# TÌM HEADER / DÒNG XÓA
# ============================================================

def find_header(ws: Any, wanted_header: str) -> tuple[int | None, int | None]:
    """Tìm header trong UsedRange; không bắt buộc ở dòng 1."""
    used = ws.UsedRange

    first_row = int(used.Row)
    first_col = int(used.Column)
    row_count = int(used.Rows.Count)
    col_count = int(used.Columns.Count)

    if row_count <= 0 or col_count <= 0:
        return None, None

    values = used.Value2

    if not isinstance(values, tuple):
        values = ((values,),)
    elif values and not isinstance(values[0], tuple):
        values = (values,)

    wanted = normalize(wanted_header)

    for r_offset, row_values in enumerate(values):
        for c_offset, value in enumerate(row_values):
            if normalize(value) == wanted:
                return first_row + r_offset, first_col + c_offset

    return None, None


def last_value_row(ws: Any, col: int, header_row: int) -> int:
    found = ws.Columns(col).Find(
        What="*",
        After=ws.Cells(header_row, col),
        LookIn=-4163,       # xlValues
        LookAt=2,           # xlPart
        SearchOrder=1,      # xlByRows
        SearchDirection=2,  # xlPrevious
        MatchCase=False,
    )

    if found is None:
        return header_row

    return int(found.Row)


def find_rows_to_delete(
    ws: Any,
    header_row: int,
    url_col: int,
) -> list[int]:
    last_row = last_value_row(ws, url_col, header_row)

    if last_row <= header_row:
        return []

    values = ws.Range(
        ws.Cells(header_row + 1, url_col),
        ws.Cells(last_row, url_col),
    ).Value2

    if not isinstance(values, tuple):
        values = ((values,),)
    elif values and not isinstance(values[0], tuple):
        values = tuple((v,) for v in values)

    wanted = {
        normalize(value)
        for value in DELETE_VALUES
    }

    rows: list[int] = []

    for row_num, row_values in enumerate(
        values,
        start=header_row + 1,
    ):
        value = (
            row_values[0]
            if isinstance(row_values, tuple)
            else row_values
        )

        if normalize(value) in wanted:
            rows.append(row_num)

    return rows


# ============================================================
# XỬ LÝ MỘT FILE BẰNG EXCEL THẬT
# ============================================================

def process_excel(excel: Any, file_path: Path) -> int:
    print()
    print(f"Đang xử lý: {file_path.name}")

    wb = None
    ws = None
    was_protected = False

    try:
        wb = excel.Workbooks.Open(
            Filename=str(file_path),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
        )

        if bool(wb.ReadOnly):
            print("  ❌ File đang ReadOnly hoặc bị khóa. Bỏ qua.")
            wb.Close(SaveChanges=False)
            return 0

        try:
            ws = wb.Worksheets(TARGET_SHEET)
        except Exception:
            print(
                f"  ⏭ Không có sheet '{TARGET_SHEET}' - bỏ qua."
            )
            wb.Close(SaveChanges=False)
            return 0

        try:
            was_protected = bool(ws.ProtectContents)
        except Exception:
            was_protected = False

        if was_protected:
            try:
                ws.Unprotect(Password=SHEET_PASSWORD)
            except Exception:
                pass

            if bool(ws.ProtectContents):
                print(
                    f"  ❌ Không mở được Protect bằng pass "
                    f"{SHEET_PASSWORD}. Bỏ qua."
                )
                wb.Close(SaveChanges=False)
                return 0

        # 1. BỎ FILTER TRƯỚC
        filters_cleared = clear_all_filters(ws)
        print(
            f"  ✓ Đã bỏ filter đang áp dụng: "
            f"{filters_cleared}"
        )

        # 2. TÌM URL PAGE
        header_row, url_col = find_header(
            ws,
            TARGET_COLUMN,
        )

        if header_row is None or url_col is None:
            print(
                f"  ⚠ Không tìm thấy cột "
                f"'{TARGET_COLUMN}'"
            )

            if was_protected:
                ws.Protect(Password=SHEET_PASSWORD)

            wb.Close(SaveChanges=False)
            return 0

        print(
            f"  ✓ Tìm thấy '{TARGET_COLUMN}' "
            f"ở dòng {header_row}, cột {url_col}"
        )

        # 3. TÌM DÒNG CẦN XÓA
        rows_to_delete = find_rows_to_delete(
            ws,
            header_row,
            url_col,
        )

        if not rows_to_delete:
            print("  ✓ Không có hàng cần xóa.")

            if was_protected:
                ws.Protect(Password=SHEET_PASSWORD)

            # Không thay đổi gì -> không Save.
            wb.Close(SaveChanges=False)
            return 0

        print(
            f"  → Có {len(rows_to_delete)} hàng cần xóa: "
            + ", ".join(str(r) for r in rows_to_delete)
        )

        # 4. XÓA NGUYÊN HÀNG TỪ DƯỚI LÊN
        for row_num in reversed(rows_to_delete):
            ws.Rows(row_num).EntireRow.Delete()

        # 5. KHÔI PHỤC PROTECT
        if was_protected:
            ws.Protect(Password=SHEET_PASSWORD)

        # 6. TÍNH LẠI CÔNG THỨC
        try:
            excel.CalculateFullRebuild()
        except Exception:
            try:
                excel.CalculateFull()
            except Exception:
                pass

        # 7. SAVE
        wb.Save()
        wb.Close(SaveChanges=False)

        print(
            f"  ✅ ĐÃ XÓA {len(rows_to_delete)} HÀNG"
        )

        return len(rows_to_delete)

    except Exception as exc:
        print(f"  ❌ Lỗi: {exc}")

        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass

        return 0


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    print("=" * 76)
    print(
        "XÓA 'BÀI ĐÃ XÓA' - "
        "ĐỌC FILE TỔNG + EXCEL COM"
    )
    print("=" * 76)

    if not ROOT_FOLDER.is_dir():
        raise RuntimeError(
            "Không tìm thấy thư mục:\n"
            + str(ROOT_FOLDER)
        )

    control_path = (
        ROOT_FOLDER / CONTROL_WORKBOOK_NAME
    )

    if not control_path.is_file():
        raise RuntimeError(
            "Không tìm thấy file tổng:\n"
            + str(control_path)
        )

    # --------------------------------------------------------
    # 1. ĐỌC FILE TỔNG
    # --------------------------------------------------------
    ready_domains, invalid_domains = (
        ready_domains_from_control_file(
            control_path
        )
    )

    if not ready_domains:
        print(
            "File tổng không có kênh nào có "
            f"PHÂN TÍCH = "
            f"'{CONTROL_ANALYSIS_COMPLETE}'."
        )
        return

    # --------------------------------------------------------
    # 2. CHỈ CHỌN FILE DOMAIN CẦN XỬ LÝ
    # --------------------------------------------------------
    files, missing, ambiguous = (
        select_ready_domain_files(
            ROOT_FOLDER,
            control_path,
            ready_domains,
        )
    )

    print()
    print(
        f"Kênh 'Phân tích xong' trong file tổng: "
        f"{len(ready_domains)}"
    )
    print(
        f"File thực tế cần mở để kiểm tra: "
        f"{len(files)}"
    )

    if files:
        print("Danh sách file:")
        for path in files:
            print(f"  - {path.name}")

    if missing:
        print()
        print(
            "⚠ Có trong file tổng nhưng chưa có file Excel:"
        )
        for domain in missing:
            print(f"  - {domain}")

    if ambiguous:
        print()
        print(
            "⚠ Có nhiều file trùng tên domain, đã bỏ qua:"
        )
        for domain in ambiguous:
            print(f"  - {domain}")

    if invalid_domains:
        print()
        print(
            "⚠ NHÓM KÊNH không phải domain hợp lệ:"
        )
        for domain in invalid_domains:
            print(f"  - {domain}")

    if not files:
        print()
        print("Không có file nào cần xử lý.")
        return

    # --------------------------------------------------------
    # 3. MỞ EXCEL RIÊNG VÀ XỬ LÝ ĐÚNG CÁC FILE ĐÃ CHỌN
    # --------------------------------------------------------
    excel = win32.DispatchEx(
        "Excel.Application"
    )

    old_screen_updating = excel.ScreenUpdating
    old_display_alerts = excel.DisplayAlerts
    old_enable_events = excel.EnableEvents

    total_files_checked = 0
    total_files_deleted = 0
    total_rows_deleted = 0

    try:
        excel.Visible = False
        excel.ScreenUpdating = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False

        for file_path in files:
            total_files_checked += 1

            deleted_count = process_excel(
                excel,
                file_path,
            )

            if deleted_count > 0:
                total_files_deleted += 1
                total_rows_deleted += deleted_count

    finally:
        try:
            excel.EnableEvents = old_enable_events
        except Exception:
            pass

        try:
            excel.DisplayAlerts = old_display_alerts
        except Exception:
            pass

        try:
            excel.ScreenUpdating = old_screen_updating
        except Exception:
            pass

        try:
            excel.Quit()
        except Exception:
            pass

    print()
    print("=" * 76)
    print("HOÀN TẤT")
    print("=" * 76)
    print(
        f"File đã kiểm tra : "
        f"{total_files_checked}"
    )
    print(
        f"File có xóa      : "
        f"{total_files_deleted}"
    )
    print(
        f"Tổng hàng đã xóa : "
        f"{total_rows_deleted}"
    )
    print("=" * 76)


if __name__ == "__main__":
    main()

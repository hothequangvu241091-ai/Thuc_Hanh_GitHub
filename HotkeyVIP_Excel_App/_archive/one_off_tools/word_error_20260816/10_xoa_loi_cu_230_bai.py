# -*- coding: utf-8 -*-
"""Chỉ xóa cột Lỗi thử lại của đúng 230 bài trong file mapping."""

import os
import sys
from pathlib import Path

import openpyxl
import xlwings as xw

ROOT = Path(__file__).resolve().parents[3]
MAPPING = ROOT / "outputs" / "worker_profile_word_errors_20260816" / "word_error_worker_mapping.xlsx"
TARGET = Path(r"D:\CodexProjects\Hotkeyvip\04_excel\hotkeyvip_test.xlsm")

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def mapping_keys():
    book = openpyxl.load_workbook(MAPPING, data_only=True)
    sheet = book["WORD_ERROR theo Worker"]
    headers = {
        str(sheet.cell(6, col).value or "").strip(): col - 1
        for col in range(1, sheet.max_column + 1)
    }
    result = set()
    for values in sheet.iter_rows(min_row=7, values_only=True):
        domain = str(values[headers["Tên Miền"]] or "").strip()
        keyword = str(values[headers["Từ khóa"]] or "").strip()
        if domain and keyword:
            result.add((domain.casefold(), keyword.casefold()))
    book.close()
    return result


def find_book():
    target = os.path.normcase(os.path.abspath(TARGET))
    for app in xw.apps:
        for book in app.books:
            if os.path.normcase(os.path.abspath(book.fullname)) == target:
                return book, False
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    return app.books.open(str(TARGET), update_links=False, read_only=False), True


def main():
    keys = mapping_keys()
    book, owns_app = find_book()
    app = book.app
    try:
        sheet = book.sheets["VIET_BAI"]
        values = sheet.used_range.value
        headers = {
            str(value or "").strip(): index
            for index, value in enumerate(values[0])
            if str(value or "").strip()
        }
        field = "Lỗi thử lại"
        changed = 0
        for row in values[1:]:
            key = (
                str(row[headers["Tên Miền"]] or "").strip().casefold(),
                str(row[headers["Từ khóa"]] or "").strip().casefold(),
            )
            if key not in keys:
                continue
            row[headers[field]] = ""
            changed += 1

        last_row = len(values)
        col = headers[field] + 1
        column_values = [[values[row - 1][headers[field]]] for row in range(2, last_row + 1)]
        sheet.range((2, col), (last_row, col)).value = column_values
        book.save()
        print(f"Đã xóa cột Lỗi thử lại: {changed} dòng / {len(keys)} bài")
    finally:
        if owns_app:
            book.close()
            app.quit()


if __name__ == "__main__":
    main()

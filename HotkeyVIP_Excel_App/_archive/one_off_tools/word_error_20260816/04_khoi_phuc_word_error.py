# -*- coding: utf-8 -*-
"""Phục hồi riêng các file Word bị WORD_ERROR từ URL ChatGPT đã có.

Không gửi prompt, không tạo ảnh, không sửa flow chính. Danh sách đầu vào là
file word_error_worker_mapping.xlsx đã đối chiếu URL với đúng Edge profile.
"""

import argparse
import importlib.util
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
import pythoncom


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


APP_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_INPUT = (
    APP_ROOT
    / "outputs"
    / "worker_profile_word_errors_20260816"
    / "word_error_worker_mapping.xlsx"
)
FLOW_PATH = APP_ROOT / "app_flows" / "03_viet_bai_tao_anh.py"
LOG_DIR = APP_ROOT / "outputs" / "word_recovery_logs"
VISIBLE_PROGRESS = Path(__file__).with_name("TIEN_DO_KHOI_PHUC_WORD.txt")
MIN_SNAPSHOT_WORDS = 80

BRIEF_REQUEST_MARKERS = (
    "===brief_1===",
    "hãy tạo 2 mô tả cảnh",
    "mô tả cảnh 1:",
    "không tạo ảnh.",
)
BRIEF_ANSWER_MARKERS = ("===brief_1===", "===brief_2===")


def load_flow_module():
    sys.path.insert(0, str(FLOW_PATH.parent))
    spec = importlib.util.spec_from_file_location("hotkeyvip_flow_word_v223", FLOW_PATH)
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def normalize_header(value):
    return str(value or "").strip()


def read_jobs(input_path, selected_workers):
    workbook = openpyxl.load_workbook(input_path, read_only=False, data_only=True)
    sheet = workbook["WORD_ERROR theo Worker"]
    header_row = None
    headers = {}
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20)):
        values = [normalize_header(cell.value) for cell in row]
        if "URL ChatGPT" in values and "Worker profile" in values:
            header_row = row[0].row
            headers = {value: index for index, value in enumerate(values) if value}
            break
    if header_row is None:
        raise RuntimeError("Không tìm thấy hàng tiêu đề trong file mapping.")

    required = [
        "STT", "Dòng Excel", "Tên Miền", "Từ khóa", "URL ChatGPT",
        "Worker profile", "Đường dẫn Word",
    ]
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError("File mapping thiếu cột: " + ", ".join(missing))

    jobs = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        worker = normalize_header(values[headers["Worker profile"]])
        url = normalize_header(values[headers["URL ChatGPT"]])
        word_path = normalize_header(values[headers["Đường dẫn Word"]])
        if not worker or not url or not word_path or worker not in selected_workers:
            continue
        jobs.append({
            "stt": values[headers["STT"]],
            "excel_row": values[headers["Dòng Excel"]],
            "web": normalize_header(values[headers["Tên Miền"]]),
            "keyword": normalize_header(values[headers["Từ khóa"]]),
            "url": url,
            "worker": worker,
            "word_path": word_path,
        })
    workbook.close()
    return jobs


def get_conversation_turns(driver):
    return driver.execute_script(
        r"""
        const result = [];
        const roleNodes = Array.from(document.querySelectorAll('[data-message-author-role]'));
        for (let index = 0; index < roleNodes.length; index += 1) {
            const roleNode = roleNodes[index];
            const turn = roleNode.closest('div[data-testid^="conversation-turn-"]') || roleNode.parentElement;
            const role = roleNode.getAttribute('data-message-author-role') || '';
            let contentNode = roleNode;
            if (role === 'assistant') {
                contentNode = roleNode.querySelector('.markdown') || (turn && turn.querySelector('.markdown')) || roleNode;
            }
            const clone = contentNode.cloneNode(true);
            clone.querySelectorAll(
                'button, svg, [data-testid*="copy"], [aria-label*="Copy"], [aria-label*="Sao chép"]'
            ).forEach(el => el.remove());
            result.push({
                index,
                role,
                text: (clone.textContent || '').trim(),
                html: role === 'assistant' ? (clone.innerHTML || '') : '',
            });
        }
        return result;
        """
    ) or []


def wait_stable_turns(driver, timeout=60, stable_seconds=2.0):
    deadline = time.time() + timeout
    previous_signature = None
    stable_since = None
    last_turns = []
    while time.time() < deadline:
        turns = get_conversation_turns(driver)
        signature = tuple(
            (turn.get("index"), turn.get("role"), turn.get("text")) for turn in turns
        )
        if turns and signature == previous_signature:
            stable_since = stable_since or time.time()
            if time.time() - stable_since >= stable_seconds:
                return turns
        else:
            previous_signature = signature
            stable_since = None
            last_turns = turns
        time.sleep(0.5)
    raise RuntimeError(f"Hội thoại chưa ổn định; chỉ đọc được {len(last_turns)} lượt.")


def contains_any(text, markers):
    lowered = str(text or "").casefold()
    return any(marker in lowered for marker in markers)


def select_article_snapshot(turns):
    """Ưu tiên assistant ngay trước prompt Brief; fallback là bài hợp lệ cuối cùng."""
    for position, turn in enumerate(turns):
        if turn.get("role") != "user":
            continue
        if not contains_any(turn.get("text"), BRIEF_REQUEST_MARKERS):
            continue
        for candidate in reversed(turns[:position]):
            if candidate.get("role") == "assistant" and candidate.get("html"):
                text = str(candidate.get("text") or "").strip()
                if not contains_any(text, BRIEF_ANSWER_MARKERS):
                    return {"text": text, "html": candidate["html"]}, "assistant_truoc_prompt_brief"

    candidates = []
    for turn in turns:
        if turn.get("role") != "assistant" or not turn.get("html"):
            continue
        text = str(turn.get("text") or "").strip()
        if contains_any(text, BRIEF_ANSWER_MARKERS):
            continue
        candidates.append((len(text.split()), turn))
    if not candidates:
        raise RuntimeError("Không tìm thấy phản hồi bài viết; chỉ thấy nội dung Brief hoặc chat rỗng.")
    _, candidate = max(candidates, key=lambda item: item[0])
    return {"text": candidate["text"], "html": candidate["html"]}, "fallback_bai_dai_nhat"


def append_checkpoint(checkpoint_path, result):
    existing = []
    if checkpoint_path.exists():
        try:
            existing = json.loads(checkpoint_path.read_text(encoding="utf-8"))
        except Exception:
            existing = []
    existing.append(result)
    checkpoint_path.write_text(
        json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def write_visible_progress(state, completed, skipped, failed, total, detail=""):
    text = (
        f"TRẠNG THÁI: {state}\n"
        f"Thành công mới: {completed}\n"
        f"Đã có sẵn, bỏ qua: {skipped}\n"
        f"Lỗi: {failed}\n"
        f"Tổng danh sách: {total}\n"
        f"Cập nhật: {datetime.now().isoformat(timespec='seconds')}\n"
        f"Chi tiết: {detail}\n"
    )
    VISIBLE_PROGRESS.write_text(text, encoding="utf-8")


def process_job(flow, driver, job, dry_run=False):
    started = time.time()
    driver.get(job["url"])
    turns = wait_stable_turns(driver)
    snapshot, method = select_article_snapshot(turns)
    word_count = flow.count_words(snapshot["text"])
    if word_count < MIN_SNAPSHOT_WORDS:
        raise RuntimeError(f"Nội dung được chọn quá ngắn: {word_count} từ.")
    if contains_any(snapshot["text"], BRIEF_ANSWER_MARKERS):
        raise RuntimeError("Nội dung được chọn vẫn chứa marker Brief; không lưu để tránh nhầm.")

    keyword_found = job["keyword"].casefold() in snapshot["text"].casefold()
    if not dry_run:
        for attempt in (1, 2):
            try:
                flow.copy_and_save_snapshot(snapshot, job["word_path"])
                break
            except Exception as exc:
                if attempt == 1 and (
                    isinstance(exc, flow.WordSystemError)
                    or flow.is_word_system_error(exc)
                ):
                    print("Word/COM lỗi lần 1; đang dọn tiến trình, preflight và thử lại bài hiện tại...")
                    if flow.recover_word_runtime_once():
                        continue
                raise
        if not flow.is_word_ok(job["word_path"]):
            raise RuntimeError("Đã lưu nhưng file Word không đạt kiểm tra đọc lại.")

    return {
        **job,
        "status": "DRY_RUN_OK" if dry_run else "OK",
        "selection": method,
        "word_count": word_count,
        "keyword_found": keyword_found,
        "seconds": round(time.time() - started, 1),
        "time": datetime.now().isoformat(timespec="seconds"),
    }


def parse_args():
    parser = argparse.ArgumentParser(description="Phục hồi Word từ URL ChatGPT đúng worker.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="File mapping Excel.")
    parser.add_argument("--workers", nargs="+", default=["worker_1", "worker_3"])
    parser.add_argument("--limit", type=int, default=0, help="0 = xử lý toàn bộ.")
    parser.add_argument("--dry-run", action="store_true", help="Chỉ kiểm tra nội dung, chưa tạo Word.")
    return parser.parse_args()


def main():
    args = parse_args()
    input_path = Path(args.input).resolve()
    selected_workers = set(args.workers)
    jobs = read_jobs(input_path, selected_workers)
    if args.limit > 0:
        jobs = jobs[: args.limit]
    if not jobs:
        raise RuntimeError("Không có bài phù hợp để phục hồi.")

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    checkpoint = LOG_DIR / f"word_recovery_{stamp}.json"
    flow = load_flow_module()

    print(f"Danh sách: {len(jobs)} bài | workers: {', '.join(sorted(selected_workers))}")
    print(f"Checkpoint: {checkpoint}")
    if not args.dry_run:
        print("Đang kiểm tra Word/VBA trước khi chạy...")
        flow.run_word_preflight("phục hồi Word lỗi")

    pythoncom.CoInitialize()
    drivers = {}
    completed = 0
    skipped = 0
    failed = 0
    write_visible_progress("ĐANG KHỞI ĐỘNG", completed, skipped, failed, len(jobs))
    try:
        for worker in sorted(selected_workers):
            worker_jobs = [job for job in jobs if job["worker"] == worker]
            if not worker_jobs:
                continue
            worker_id = int(worker.rsplit("_", 1)[1])
            flow._THREAD_CONTEXT.worker_id = worker_id
            print(f"\n===== {worker}: {len(worker_jobs)} bài =====")
            driver, _ = flow.create_shared_driver()
            drivers[worker] = driver

            for position, job in enumerate(worker_jobs, start=1):
                print(
                    f"[{worker} {position}/{len(worker_jobs)}] "
                    f"Dòng {job['excel_row']}: {job['keyword']}"
                )
                try:
                    if not args.dry_run and flow.is_word_ok(job["word_path"]):
                        skipped += 1
                        result = {
                            **job,
                            "status": "ALREADY_OK",
                            "time": datetime.now().isoformat(timespec="seconds"),
                        }
                        append_checkpoint(checkpoint, result)
                        write_visible_progress(
                            "ĐANG CHẠY", completed, skipped, failed, len(jobs),
                            f"{worker} {position}/{len(worker_jobs)} | đã có Word: {job['keyword']}",
                        )
                        print("BỎ QUA: file Word đã có và đọc được.")
                        continue
                    result = process_job(flow, driver, job, dry_run=args.dry_run)
                    append_checkpoint(checkpoint, result)
                    completed += 1
                    warning = "" if result["keyword_found"] else " | cảnh báo: không thấy từ khóa nguyên văn"
                    print(f"OK: {result['word_count']} từ | {result['selection']}{warning}")
                    write_visible_progress(
                        "ĐANG CHẠY", completed, skipped, failed, len(jobs),
                        f"{worker} {position}/{len(worker_jobs)} | vừa xong: {job['keyword']}",
                    )
                except Exception as exc:
                    failed += 1
                    result = {
                        **job,
                        "status": "ERROR",
                        "error": str(exc),
                        "time": datetime.now().isoformat(timespec="seconds"),
                    }
                    append_checkpoint(checkpoint, result)
                    print(f"LỖI: {exc}")
                    write_visible_progress(
                        "GẶP LỖI", completed, skipped, failed, len(jobs),
                        f"{worker} {position}/{len(worker_jobs)} | {job['keyword']} | {exc}",
                    )
                    if isinstance(exc, flow.WordSystemError) or flow.is_word_system_error(exc):
                        try:
                            import winsound
                            for _ in range(5):
                                winsound.Beep(1200, 350)
                        except Exception:
                            pass
                        raise RuntimeError(
                            "Word/COM lỗi hệ thống. Đã dừng toàn bộ phục hồi để không sinh thêm Word lỗi."
                        ) from exc
    finally:
        for driver in drivers.values():
            try:
                driver.quit()
            except Exception:
                pass
        flow.cleanup_all_owned_word_processes()
        pythoncom.CoUninitialize()

    write_visible_progress(
        "HOÀN TẤT", completed, skipped, failed, len(jobs), "Đã xử lý hết danh sách.",
    )
    print(f"\nHOÀN TẤT: {completed} thành công mới | {skipped} đã có | {failed} lỗi")
    print(f"Nhật ký: {checkpoint}")


if __name__ == "__main__":
    main()
